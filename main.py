import json
import os
import shutil
import time
from collections import OrderedDict
import openpyxl as xl
from openpyxl.utils import column_index_from_string
from fastapi import FastAPI, Request, WebSocket, HTTPException, Body
from typing import List
from fastapi import Form
from fastapi.responses import FileResponse
from fastapi.responses import HTMLResponse
from fastapi.responses import RedirectResponse
from fastapi.responses import Response
from fastapi.security import HTTPBasic
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from fastapi import File, UploadFile
from fastapi.responses import JSONResponse
from utils.Folder_usual import creat_folder, Save_upload_file, zip_unzip, creat_function_folder
from utils.progerss_excel import get_file_path, fujian4, create_fujian3, create_fujian2, create_fujian1, \
    creat_info_project_excel
from utils.progerss_excel import fb_excel, dw_excel, zj_excel, dj_excel, get_project_data_dict
from utils.jdk_xd import jdk_xd
from docx.shared import Pt
from docx.oxml.ns import qn
from docx import Document
import datetime
import re
from pydantic import BaseModel

app = FastAPI()
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")


def read_data(filename):
    filepath = os.path.join(os.path.dirname(__file__), 'static', 'res')
    path = os.path.join(filepath, f'{filename}.json')
    with open(path, 'r', encoding='utf-8') as f:
        result = json.load(f)
    return result


def save_data(filename, data):
    filepath = os.path.join(os.path.dirname(__file__), 'static', 'res')
    path = os.path.join(filepath, f'{filename}.txt')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(str(data))
        f.close()
    return path


@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/{page}.html", response_class=HTMLResponse)
async def page(request: Request, page: str):
    return templates.TemplateResponse(f'{page}.html', {"request": request})


@app.get("/library", response_class=HTMLResponse)
async def library(request: Request):
    user_name = request.session.get('username')
    return templates.TemplateResponse("price_content.html", {"request": request})


def getLibraryPlantsNameList():
    file_path = os.path.join(os.path.dirname(__file__), 'resources', '苗木价格.xlsx')
    wb = xl.load_workbook(file_path, data_only=True)
    ws = wb['常规苗木 (2)']
    name_list = []

    for row in ws.iter_rows():
        if row[0].row < 4:
            continue
        if row[column_index_from_string("C") - 1].value is None:
            continue
        name_list.append(row[column_index_from_string("B") - 1].value)
        name_list = list(OrderedDict.fromkeys(name_list))
    file_name = '秀林苗木名称列表'
    file_path = save_data(filename=file_name, data=name_list)
    return file_path


def getLibraryPlantsData():
    file_path = os.path.join(os.path.dirname(__file__), 'resources', '苗木价格.xlsx')
    wb = xl.load_workbook(file_path, data_only=True)
    ws = wb['常规苗木 (2)']
    result_list = []
    for row in ws.iter_rows():
        if row[0].row < 4:
            continue
        if not row[1].value:
            continue
        a_dict = {
            "xuhao": row[column_index_from_string("A") - 1].value,
            "mingcheng": row[column_index_from_string("B") - 1].value,
            "zhonglei": row[column_index_from_string("C") - 1].value,
            "guige_gaodu": row[column_index_from_string("D") - 1].value,
            "guige_xiongdijing": row[column_index_from_string("E") - 1].value,
            "guige_guanfu": row[column_index_from_string("F") - 1].value,
            "guige_fenzhidian": row[column_index_from_string("G") - 1].value,
            "danwei": row[column_index_from_string("H") - 1].value,
            "daochangjia_buhanshui": round(row[column_index_from_string("J") - 1].value, 2)
        }
        result_list.append(a_dict)
    file_name = '秀林苗木价格字典'
    file_path = save_data(file_name, result_list)
    return file_path


@app.get("/search_data_index", response_class=HTMLResponse)  # 默认材料价格库显示界面
async def search_data_index(request: Request):
    try:
        name_list = read_data('秀林苗木名称列表')
    except:
        # 更新苗木名称列表
        getLibraryPlantsNameList()
        # 更新苗木库
        getLibraryPlantsData()
        name_list = read_data('秀林苗木名称列表')
    # ic(name_list)
    return templates.TemplateResponse("library.html", {"request": request, "name_list": name_list})


@app.get('/refresh_library_data', response_class=HTMLResponse)
async def refresh_library_data(request: Request):
    # 更新苗木名称列表
    getLibraryPlantsNameList()
    # 更新苗木库
    getLibraryPlantsData()
    content = '''
    <script>alert('已完成更新！');window.location.href='/search_data_index'</script>
    '''
    return HTMLResponse(content)


@app.post("/search_data", response_class=HTMLResponse)
async def search_data(request: Request, name: str = Form(...)):
    result = []
    price_data = read_data('秀林苗木价格字典')
    for value in price_data:
        if name.lower() in value['mingcheng'].lower():
            result.append(value)
    name_list = read_data('秀林苗木名称列表')

    if not result:
        return "<script>alert('未查询到数据！');window.location.href='search_data_index'</script>"
    else:
        return templates.TemplateResponse("library.html", {"request": request,
                                                           "name_list": name_list, "result": result})

#
# @app.get("/jiagequshi", response_class=HTMLResponse)
# async def root(request: Request):
#     from utils.get_data import get_name_list
#     # 依次返回 常绿、落叶、灌木、地被类 四个列表
#     changlv_name_list = get_name_list()[1]
#     luoye_name_list = get_name_list()[2]
#     guanmu_name_list = get_name_list()[3]
#     dibeilei_name_list = get_name_list()[4]
#     fig_html = "暂无数据"
#     return templates.TemplateResponse("qushi.html", {"request": request,  # request参数必须有
#                                                      "changlv_name_list": changlv_name_list,  # 常绿列表
#                                                      "luoye_name_list": luoye_name_list,  # 落叶列表
#                                                      "guanmu_name_list": guanmu_name_list,  # 灌木列表
#                                                      "dibeilei_name_list": dibeilei_name_list})  # 地被类列表
#
# # 后台手动刷新数据
# @app.get("/f", response_class=HTMLResponse)  # 用于前端ajax请求
# async def refresh_data(request: Request):
#     from utils.get_data import get_data
#     from utils.get_data import get_name_list
#     from utils.get_data import draw
#     name_list = get_name_list()[0]
#     dir = "static/result/"
#     try:
#         os.makedirs(dir)
#     except FileExistsError:
#         pass
#     for name in name_list:
#         fig = draw(get_data(name)[0], get_data(name)[1], name)
#         fig = fig.write_html(f"static/result/{name}.html")
#     html_text = ("<script>"
#                  "alert('数据刷新成功');"
#                  "window.location.href='/jiagequshi';"
#                  "</script>")
#     return HTMLResponse(content=html_text, status_code=200)


@app.get('/xinxijia')
def xinxijia(request: Request):
    return templates.TemplateResponse("信息价查询.html", {"request": request})


@app.post("/xinxijia_check", response_class=HTMLResponse)
async def xinxijia_check(request: Request, sj: str = Form(default=datetime.datetime.now().strftime('%Y-%m-%d')),
                         mc: str = Form(default=""), gg: str = Form(default="")):
    print(sj)
    shijian_year = sj.split('-')[0]
    shijian_month = sj.split('-')[1].replace('0', '')
    shijian = f'{shijian_year}年{shijian_month}月'
    result = []
    starttime = datetime.datetime.now()
    price_data = read_data('雄安新区信息价')
    for value in price_data:
        if mc.lower() in value['mc'].lower() and shijian == value['sj'] and gg in value['gg']:
            result.append(value)
    endtime = datetime.datetime.now()
    haoshi = (endtime - starttime).seconds
    if not result:
        result = [{
            'sj': '没',
            'mc': '找',
            'gg': '到，',
            'dw': '对',
            'p': '不',
            'p_tax': '起！',
        }]
    return templates.TemplateResponse("信息价查询.html", {"request": request, "datelist": result, "haoshi": haoshi})


@app.get("/swift_get_data")
async def swift_get_data(request: Request):
    result = []
    price_data = read_data(filename='秀林苗木价格字典')
    for value in price_data:
        mc = str(value['mingcheng'])
        zl = str(value['zhonglei'])
        gd = str(value['guige_gaodu'])
        xdj = str(value['guige_xiongdijing'])
        gg = str(value['guige_guanfu'])
        fzd = str(value['guige_fenzhidian'])
        dw = str(value['danwei'])
        jg = str(value['daochangjia_buhanshui'])
        dataList = [mc, zl, gd, xdj, gg, fzd, dw, jg]
        result.append(dataList)
    return result


@app.post("/shortcut/oil")
async def shortcut_oil(request: Request):
    message = await request.json()
    gonglishu = message['gls']
    youjia = message['yj']
    jiayouliang = message['jyl']
    print(message)
    return gonglishu


@app.post("/shortcut/zhangben")
async def shortcut_zhangben(request: Request):
    message = await request.json()
    xinxi = message['xinxi']
    print(xinxi)
    xinxi = '【北京银行】您账户1160于5月21日20:31通过网银在线（京东支付）支付59.40元。活期余额603.13元。对方户名:京东商城业务。'

    def exctract_info(sms):
        account_pattern = r'账户(\d+)'
        # date_time_pattern = r'\d{4}-\d{2}-\d{2}'
        amount_pattern = r'支付([\d.]+)元'
        account = re.search(account_pattern, sms).group(1)
        # date_time = re.search(date_time_pattern, sms).group(1)
        amount = re.search(amount_pattern, sms).group(1)
        return {
            'account': account,
            # 'date_time': date_time,
            'amount': amount
        }

    info = exctract_info(xinxi)
    print(info)

    return info


@app.get("/ceshi_html", response_class=HTMLResponse)
async def ceshi_html(request: Request):
    # 获取列表
    filename = os.path.join(os.path.dirname(__file__), 'static', 'res', '2024年3月雄安信息价字典.json')
    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)
    result_key = []
    result_value = []
    valueList = []
    for key, value in data.items():
        result_key.append(key)
        result_value.append(value)
    for value in result_value:
        for v in value:
            valueList.append(v)

    return templates.TemplateResponse("ceshi.html", {"request": request, "keylist": result_key, "valuelist": valueList})


# 进度款
@app.get("/creat_reports_zip_first", response_class=HTMLResponse)
async def creat_reports_zip_first(request: Request):
    try:
        username = request.session.get('username')
    except:
        username = 'Guest'
    print(username)
    return templates.TemplateResponse("Quick_progress_report_xd_first_time_3_4.html", {"request": request})


@app.post("/creat_reports_xd_result_3_4", response_class=HTMLResponse)
async def creat_reports_xd_result(request: Request,
                                  Contract_Name: str = Form(default="合同名称"),  # 合同名称
                                  Contract_Number: str = Form(default="合同编号"),  # 合同编号
                                  Principal_Party: str = Form(default="建设单位"),  # 甲方
                                  Construction_Unit: str = Form(default="施工单位"),  # 施工单位
                                  Supervision_Unit: str = Form(default="监理单位"),  # 监理单位
                                  Design_Unit: str = Form(default="设计单位"),  # 设计单位
                                  Consulting_Unit: str = Form(default="咨询单位"),  # 咨询单位
                                  Progress_Payment_Ratio=Form(default=0.85),  # 进度款比例
                                  Contract_Amount=Form(default=0),  # 合同金额
                                  Reporting_Periods: str = Form(default=1),  # 报告期数
                                  Funding_Source: str = Form(default="新区财政资金"),  # 资金来源
                                  Price_Form: str = Form(default="固定单价合同"),  # 价格形式
                                  Advance_Payment_Payment_Ratio: str = Form(default=0),  # 预付款支付比例
                                  Advance_Payment_Deduction_Ratio: str = Form(default=0),  # 预付款扣回比例
                                  Project_Overview: str = Form(default="暂无"),  # 工程概况
                                  file_input_ht: UploadFile = File(default=None),
                                  file_input_ss: UploadFile = File(default=None),
                                  file_input_sd: UploadFile = File(default=None)
                                  ):
    try:
        username = request.session.get('username')
    except:
        username = 'Guest'
    global function_dir
    function_dir = os.path.join(os.path.dirname(__file__), 'user_folder', f'{username}')
    print(f"function_dir:{function_dir}")
    start_time = time.time()
    year = datetime.datetime.now().strftime("%Y")
    month = datetime.datetime.now().strftime("%m")
    # 创建一个用于保存主程序过程文件的文件夹，我将它命名为progress_reports_xd，意味对接新点软件的进度报告
    project_dir = os.path.join(os.path.dirname(__file__), "user_folder", f'{username}',
                               "progress_reports_xd", Contract_Name, f'第{Reporting_Periods}期')
    print(project_dir)
    try:
        os.makedirs(project_dir)
    except FileExistsError:
        print(f'{project_dir}已存在')
        shutil.rmtree(project_dir)
        os.makedirs(project_dir, exist_ok=True)

    # 对于本项目，创建一个临时用Excel文档，用于保存历史记录，下次用户只需要再次输入名称，则可以调取相应的信息内容
    progress_reports_xd_dir = os.path.join(function_dir, "progress_reports_xd")
    print('progress_reports_xd_dir', progress_reports_xd_dir)
    project_info_dict = creat_info_project_excel(
        Contract_Name=Contract_Name,
        output_path=progress_reports_xd_dir,
        Contract_Number=Contract_Number,
        Principal_Party=Principal_Party,
        Construction_Unit=Construction_Unit,
        Supervision_Unit=Supervision_Unit,
        Design_Unit=Design_Unit,
        Consulting_Unit=Consulting_Unit,
        Project_Overview=Project_Overview,
        Funding_Source=Funding_Source,
        Price_Form=Price_Form,
        Progress_Payment_Ratio=Progress_Payment_Ratio,
        Contract_Amount=Contract_Amount,
        Advance_Payment_Payment_Ratio=Advance_Payment_Payment_Ratio,
        Advance_Payment_Deduction_Ratio=Advance_Payment_Deduction_Ratio
    )
    input_folder_path = creat_folder(os.path.join(project_dir, "input_files"))  # 输入文件路径
    output_folder_path = creat_folder(os.path.join(project_dir, "output_files"))  # 输出文件路径
    process_folder_path = creat_folder(os.path.join(project_dir, "process_files"))  # 过程文件路径

    ht_folder = creat_folder(os.path.join(input_folder_path, "ht"))  # 新建合同文件夹
    send_folder = creat_folder(os.path.join(input_folder_path, "ss"))  # 新建送审文件夹
    audited_folder = creat_folder(os.path.join(input_folder_path, "sd"))  # 新建审定文件夹

    # 解压合同文件至过程文件夹
    ht_zip = Save_upload_file(file_input_ht, ht_folder, f"合同.zip")
    # print(ht_zip)
    zip_unzip(zip_file_name=ht_zip, target_dir=ht_folder)
    send_zip = Save_upload_file(file_input_ss, send_folder, f"送审.zip")
    # print(send_zip)
    zip_unzip(zip_file_name=send_zip, target_dir=send_folder)
    audited_zip = Save_upload_file(file_input_sd, audited_folder, f"审定.zip")
    # print(audited_zip)
    zip_unzip(zip_file_name=audited_zip, target_dir=audited_folder)

    # 处理表格
    get_file_path(ht_folder, process_folder_path, 'ht')
    get_file_path(send_folder, process_folder_path, 'ss')
    get_file_path(audited_folder, process_folder_path, 'sd')

    # 定义过程-ht拆分文件夹路径
    ht_excel_file_path_fb = creat_folder(os.path.join(process_folder_path, "ht", 'fb'))
    ht_excel_file_path_dw = creat_folder(os.path.join(process_folder_path, "ht", 'dw'))
    ht_excel_file_path_dj = creat_folder(os.path.join(process_folder_path, "ht", 'dj'))
    ht_excel_file_path_zj = creat_folder(os.path.join(process_folder_path, "ht", 'zj'))
    send_excel_file_path_fb = creat_folder(os.path.join(process_folder_path, "ss", 'fb'))
    send_excel_file_path_dw = creat_folder(os.path.join(process_folder_path, "ss", 'dw'))
    send_excel_file_path_dj = creat_folder(os.path.join(process_folder_path, "ss", 'dj'))
    send_excel_file_path_zj = creat_folder(os.path.join(process_folder_path, "ss", 'zj'))
    audited_excel_file_path_fb = creat_folder(os.path.join(process_folder_path, "sd", 'fb'))
    audited_excel_file_path_dw = creat_folder(os.path.join(process_folder_path, "sd", 'dw'))
    audited_excel_file_path_dj = creat_folder(os.path.join(process_folder_path, "sd", 'dj'))
    audited_excel_file_path_zj = creat_folder(os.path.join(process_folder_path, "sd", 'zj'))

    # 二次处理表格（删行）
    for leixing in ["ht", "ss", "sd"]:
        leixing_path = os.path.join(process_folder_path, leixing)
        for file in os.listdir(leixing_path):
            excel_path = os.path.join(leixing_path, file)
            # print(excel_path)
            if file == "fb":
                fb_excel(excel_path)
            elif file == "dj":
                dj_excel(excel_path)
            elif file == "dw":
                dw_excel(excel_path)
            else:
                zj_excel(excel_path)

    fujian4_save_path = fujian4(
        Contract_Name=Contract_Name, Reporting_Periods=Reporting_Periods, year=year, month=month,
        Contract_Number=Contract_Number,
        fenbu_path_ht=ht_excel_file_path_fb, djcs_path_ht=ht_excel_file_path_dj,
        zjcs_path_ht=ht_excel_file_path_zj, danwei_path_ht=ht_excel_file_path_dw,
        fenbu_path_sent=send_excel_file_path_fb, djcs_path_sent=send_excel_file_path_dj,
        zjcs_path_sent=send_excel_file_path_zj, danwei_path_sent=send_excel_file_path_dw,
        fenbu_path_audited=audited_excel_file_path_fb,
        djcs_path_audited=audited_excel_file_path_dj,
        zjcs_path_audited=audited_excel_file_path_zj,
        danwei_path_audited=audited_excel_file_path_dw,
        output_path=output_folder_path, input_ht_dx=ht_folder
    )
    print(f"附件4已生成，完成时间为{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    fujian3_save_path = create_fujian3(
        input_path=input_folder_path,
        output_path=output_folder_path,
        Contract_Name=Contract_Name,
        Contract_Number=Contract_Number,
        Reporting_Periods=Reporting_Periods,
        year=year, month=month
    )
    print(f"附件3已生成，完成时间为{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    # 强制删除project_dir文件夹
    shutil.rmtree(os.path.dirname(os.path.dirname(ht_excel_file_path_fb)))
    print("过程文件已删除")
    end_time = time.time()
    print(f"已完成操作，整体耗时{round(end_time - start_time, 2)}")
    print(f'合同编号是：{Contract_Number},期数是：{Reporting_Periods},年份是：{year},月份是：{month}')

    wb_fujian3 = xl.load_workbook(fujian3_save_path)
    ws_fujian3 = wb_fujian3.active
    Current_Application_Internal_Contract_Project = round(ws_fujian3[f'E{ws_fujian3.max_row}'].value, 2)
    Current_Approved_Internal_Contract_Project = round(ws_fujian3[f'F{ws_fujian3.max_row}'].value, 2)

    zip_file_name = f"{Contract_Name}_{Reporting_Periods}_{year}_{month}"
    zip_out_path = os.path.join(os.path.dirname(output_folder_path), zip_file_name)
    shutil.make_archive(zip_out_path, "zip", output_folder_path)
    zip_dir = zip_out_path + ".zip"
    zip_name = zip_file_name + ".zip"

    Current_Application_Prepayment_Safety_Civilized_Construction_Fee = 0
    Current_Application_Change_Project = 0
    Current_Application_Time_and_Materials_Project = 0
    Current_Application_Claim_Project = 0
    Current_Application_Price_Adjustment = 0
    Current_Application_Other = 0
    Current_Application_Tax_Adjustment_Payable_Amount = 0
    Current_Application_Financial_Evaluation_Reduction = 0
    Current_Application_Total_Completed_Output_Value = Current_Application_Internal_Contract_Project + Current_Application_Prepayment_Safety_Civilized_Construction_Fee + \
                                                       Current_Application_Change_Project + Current_Application_Time_and_Materials_Project + Current_Application_Claim_Project + \
                                                       Current_Application_Price_Adjustment + Current_Application_Price_Adjustment + Current_Application_Other + \
                                                       Current_Application_Tax_Adjustment_Payable_Amount - Current_Application_Financial_Evaluation_Reduction
    Current_Approved_Prepayment_Safety_Civilized_Construction_Fee = 0
    Current_Approved_Change_Project = 0
    Current_Approved_Time_and_Materials_Project = 0
    Current_Approved_Claim_Project = 0
    Current_Approved_Price_Adjustment = 0
    Current_Approved_Other = 0
    Current_Approved_Financial_Evaluation_Approval = 0
    Current_Approved_Tax_Adjustment_Payable_Amount = 0
    Current_Approved_Total_Completed_Output_Value = Current_Approved_Internal_Contract_Project + Current_Approved_Prepayment_Safety_Civilized_Construction_Fee + \
                                                    Current_Approved_Change_Project + Current_Approved_Time_and_Materials_Project + \
                                                    Current_Approved_Time_and_Materials_Project + Current_Approved_Claim_Project + \
                                                    Current_Approved_Price_Adjustment + Current_Approved_Price_Adjustment + Current_Approved_Other + \
                                                    Current_Approved_Financial_Evaluation_Approval - Current_Approved_Tax_Adjustment_Payable_Amount

    html_response = templates.TemplateResponse("Quick_progress_report_xd_first_time_1_2.html",
                                               {
                                                   "request": request,
                                                   "Contract_Name": Contract_Name,
                                                   "Contract_Number": Contract_Number,
                                                   "Principal_Party": Principal_Party,
                                                   "Construction_Unit": Construction_Unit,
                                                   "Supervision_Unit": Supervision_Unit,
                                                   "Design_Unit": Design_Unit,
                                                   "Consulting_Unit": Consulting_Unit,
                                                   "Reporting_Periods": Reporting_Periods,
                                                   "Funding_Source": Funding_Source,
                                                   "Price_Form": Price_Form,
                                                   "Project_Overview": Project_Overview,
                                                   "Contract_Amount": Contract_Amount,
                                                   "Progress_Payment_Ratio": Progress_Payment_Ratio,
                                                   "Current_Application_Internal_Contract_Project": Current_Application_Internal_Contract_Project,
                                                   "Current_Application_Prepayment_Safety_Civilized_Construction_Fee": Current_Application_Prepayment_Safety_Civilized_Construction_Fee,
                                                   "Current_Application_Change_Project": Current_Application_Change_Project,
                                                   "Current_Application_Time_and_Materials_Project": Current_Application_Time_and_Materials_Project,
                                                   "Current_Application_Claim_Project": Current_Application_Claim_Project,
                                                   "Current_Application_Price_Adjustment": Current_Application_Price_Adjustment,
                                                   "Current_Application_Other": Current_Application_Other,
                                                   "Current_Application_Tax_Adjustment_Payable_Amount": Current_Application_Tax_Adjustment_Payable_Amount,
                                                   "Current_Application_Financial_Evaluation_Reduction": Current_Application_Financial_Evaluation_Reduction,
                                                   "Current_Application_Total_Completed_Output_Value": Current_Application_Total_Completed_Output_Value,
                                                   "Current_Approved_Internal_Contract_Project": Current_Approved_Internal_Contract_Project,
                                                   "Current_Approved_Prepayment_Safety_Civilized_Construction_Fee": Current_Approved_Prepayment_Safety_Civilized_Construction_Fee,
                                                   "Current_Approved_Change_Project": Current_Approved_Change_Project,
                                                   "Current_Approved_Time_and_Materials_Project": Current_Approved_Time_and_Materials_Project,
                                                   "Current_Approved_Claim_Project": Current_Approved_Claim_Project,
                                                   "Current_Approved_Price_Adjustment": Current_Approved_Price_Adjustment,
                                                   "Current_Approved_Other": Current_Approved_Other,
                                                   "Current_Approved_Financial_Evaluation_Approval": Current_Approved_Financial_Evaluation_Approval,
                                                   "Current_Approved_Tax_Adjustment_Payable_Amount": Current_Approved_Tax_Adjustment_Payable_Amount,
                                                   "Current_Approved_Total_Completed_Output_Value": Current_Approved_Total_Completed_Output_Value,
                                                   "year": year,
                                                   "month": month,
                                                   "zip_dir": zip_dir,
                                                   "zip_name": zip_name
                                               }
                                               )
    end_time = time.time()
    print(f"已完成操作，整体耗时{round(end_time - start_time, 2)}")
    return html_response


@app.post("/fj_3_4_download", response_class=FileResponse)
async def fj_3_4_download(request: Request, zip_dir=Form(...), zip_name=Form(...)):
    print(zip_dir, zip_name)
    return FileResponse(path=zip_dir, filename=zip_name)


@app.post("/creat_reports_xd_result_1_2", response_class=HTMLResponse)
async def creat_reports_xd_result(request: Request,
                                  Contract_Name=Form(...),
                                  Reporting_Periods=Form(...),
                                  year=Form(...),
                                  month=Form(...),
                                  Completed_Contract_Internal_Projects_Until_Last_Period: str = Form(default=0),
                                  Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period: str = Form(
                                      default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Change_Projects: str = Form(default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project: str = Form(
                                      default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Claim_Project: str = Form(default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment: str = Form(
                                      default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Other: str = Form(default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount: str = Form(
                                      default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction: str = Form(
                                      default=0),
                                  Total_Completed_Output_Value: str = Form(default=0),
                                  Deduction_of_Supply_of_Plants_by_Party_A: str = Form(default=0),
                                  Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A: str = Form(
                                      default=0),
                                  Accounts_Payable_Advance_Payment: str = Form(default=0),
                                  Advance_Payment_Offset: str = Form(default=0),
                                  Temporary_Withholding_Payment_Retention_Money: str = Form(default=0),
                                  Refund_of_Payment_Retention_Money: str = Form(default=0),
                                  Total_Amount_Receivable: str = Form(default=0),
                                  Deduction_Penalty: str = Form(default=0),
                                  Deduction_Fine: str = Form(default=0),
                                  Deduction_Other: str = Form(default=0),
                                  Total_Deductions: str = Form(default=0),
                                  Total_Accounts_Payable: str = Form(default=0),
                                  # 本次送审的参数
                                  Current_Application_Internal_Contract_Project: str = Form(default=0),  #
                                  Current_Application_Prepayment_Safety_Civilized_Construction_Fee: str = Form(
                                      default=0),
                                  Current_Application_Change_Project: str = Form(default=0),
                                  Current_Application_Time_and_Materials_Project: str = Form(default=0),
                                  Current_Application_Claim_Project: str = Form(default=0),
                                  Current_Application_Price_Adjustment: str = Form(default=0),
                                  Current_Application_Other: str = Form(default=0),
                                  Current_Application_Tax_Adjustment_Payable_Amount: str = Form(default=0),
                                  Current_Application_Financial_Evaluation_Reduction: str = Form(default=0),
                                  Current_Application_Total_Completed_Output_Value: str = Form(default=0),
                                  Current_Application_Deduction_of_Supply_of_Plants_by_Party_A: str = Form(default=0),
                                  Current_Application_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A: str = Form(
                                      default=0),
                                  Current_Application_Accounts_Payable_Advance_Payment: str = Form(default=0),
                                  Current_Application_Advance_Payment_Offset: str = Form(default=0),
                                  Current_Application_Temporary_Withholding_Payment_Retention_Money: str = Form(
                                      default=0),
                                  Current_Application_Refund_of_Payment_Retention_Money: str = Form(default=0),
                                  Current_Application_Total_Amount_Receivable: str = Form(default=0),
                                  Current_Application_Deduction_Penalty: str = Form(default=0),
                                  Current_Application_Deduction_Fine: str = Form(default=0),
                                  Current_Application_Deduction_Other: str = Form(default=0),
                                  Current_Application_Total_Deductions: str = Form(default=0),
                                  Current_Application_Total_Accounts_Payable: str = Form(default=0),
                                  # 本次审核参数
                                  Current_Approved_Internal_Contract_Project: str = Form(default=0),
                                  Current_Approved_Prepayment_Safety_Civilized_Construction_Fee: str = Form(default=0),
                                  Current_Approved_Change_Project: str = Form(default=0),
                                  Current_Approved_Time_and_Materials_Project: str = Form(default=0),
                                  Current_Approved_Claim_Project: str = Form(default=0),
                                  Current_Approved_Price_Adjustment: str = Form(default=0),
                                  Current_Approved_Other: str = Form(default=0),
                                  Current_Approved_Tax_Adjustment_Payable_Amount: str = Form(default=0),
                                  Current_Approved_Financial_Evaluation_Approval: str = Form(default=0),
                                  Current_Approved_Total_Completed_Output_Value: str = Form(default=0),
                                  Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A: str = Form(default=0),
                                  Current_Approved_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A: str = Form(
                                      default=0),
                                  Current_Approved_Accounts_Payable_Advance_Payment: str = Form(default=0),
                                  Current_Approved_Advance_Payment_Offset: str = Form(default=0),
                                  Current_Approved_Temporary_Withholding_Payment_Retention_Money: str = Form(default=0),
                                  Current_Approved_Refund_of_Payment_Retention_Money: str = Form(default=0),
                                  Current_Approved_Total_Amount_Receivable: str = Form(default=0),
                                  Current_Approved_Deduction_Penalty: str = Form(default=0),
                                  Current_Approved_Deduction_Fine: str = Form(default=0),
                                  Current_Approved_Deduction_Other: str = Form(default=0),
                                  Current_Approved_Total_Deductions: str = Form(default=0),
                                  Current_Approved_Total_Accounts_Payable: str = Form(default=0),
                                  fj3: UploadFile = File(default=None),
                                  fj4: UploadFile = File(default=None)):
    start_time = time.time()
    try:
        username = request.session.get('username')
    except:
        username = 'Guest'
    print(year, month, Contract_Name, Reporting_Periods)
    project_dir = os.path.join(os.path.dirname(__file__), "user_folder", f'{username}',
                               "progress_reports_xd", Contract_Name, f'第{Reporting_Periods}期')
    project_info_xlsx_dir = os.path.join(os.path.dirname(__file__), "user_folder", f'{username}', "progress_reports_xd")
    output_folder_path = creat_folder(os.path.join(project_dir, "output_files"))  # 输出文件路径

    project_info_dict = get_project_data_dict(output_path=project_info_xlsx_dir, Contract_Name=Contract_Name)

    fujian2_save_path = create_fujian2(
        output_path=output_folder_path,
        Contract_Name=Contract_Name, Contract_Number=project_info_dict['Contract_Number'],
        Contract_Amount=project_info_dict['Contract_Amount'], year=year, month=month,
        Reporting_Periods=Reporting_Periods,
        Advance_Payment_Payment_Ratio=project_info_dict['Advance_Payment_Payment_Ratio'],
        Advance_Payment_Deduction_Ratio=project_info_dict['Advance_Payment_Deduction_Ratio'],
        Completed_Contract_Internal_Projects_Until_Last_Period=Completed_Contract_Internal_Projects_Until_Last_Period,
        Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period=Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period,
        Accumulated_Completed_Amount_Until_Last_Period_Change_Projects=Accumulated_Completed_Amount_Until_Last_Period_Change_Projects,
        Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project=Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project,
        # 计日工项目
        Accumulated_Completed_Amount_Until_Last_Period_Claim_Project=Accumulated_Completed_Amount_Until_Last_Period_Claim_Project,
        # 索赔项目
        Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment=Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment,
        Accumulated_Completed_Amount_Until_Last_Period_Other=Accumulated_Completed_Amount_Until_Last_Period_Other,
        Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount=Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount,
        Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction=Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction,
        Total_Completed_Output_Value=Total_Completed_Output_Value,
        Deduction_of_Supply_of_Plants_by_Party_A=Deduction_of_Supply_of_Plants_by_Party_A,
        Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A=Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
        Accounts_Payable_Advance_Payment=Accounts_Payable_Advance_Payment,
        Advance_Payment_Offset=Advance_Payment_Offset,
        Temporary_Withholding_Payment_Retention_Money=Temporary_Withholding_Payment_Retention_Money,
        Refund_of_Payment_Retention_Money=Refund_of_Payment_Retention_Money,  # 返还支付保留金
        Total_Amount_Receivable=Total_Amount_Receivable,
        Deduction_Penalty=Deduction_Penalty,
        Deduction_Fine=Deduction_Fine,
        Deduction_Other=Deduction_Other,
        Total_Deductions=Total_Deductions,
        Total_Accounts_Payable=Total_Accounts_Payable,
        Current_Application_Internal_Contract_Project=Current_Application_Internal_Contract_Project,
        Current_Application_Prepayment_Safety_Civilized_Construction_Fee=Current_Application_Prepayment_Safety_Civilized_Construction_Fee,
        # 预付安全文明施工费
        Current_Application_Change_Project=Current_Application_Change_Project,  # 变更项目
        Current_Application_Time_and_Materials_Project=Current_Application_Time_and_Materials_Project,  # 计日工项目
        Current_Application_Claim_Project=Current_Application_Claim_Project,  # 索赔项目
        Current_Application_Price_Adjustment=Current_Application_Price_Adjustment,
        Current_Application_Other=Current_Application_Other,
        Current_Application_Tax_Adjustment_Payable_Amount=Current_Application_Tax_Adjustment_Payable_Amount,
        Current_Application_Financial_Evaluation_Reduction=Current_Application_Financial_Evaluation_Reduction,
        Current_Application_Total_Completed_Output_Value=Current_Application_Total_Completed_Output_Value,
        Current_Application_Deduction_of_Supply_of_Plants_by_Party_A=Current_Application_Deduction_of_Supply_of_Plants_by_Party_A,
        Current_Application_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A=Current_Application_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
        Current_Application_Accounts_Payable_Advance_Payment=Current_Application_Accounts_Payable_Advance_Payment,
        Current_Application_Advance_Payment_Offset=Current_Application_Advance_Payment_Offset,
        Current_Application_Temporary_Withholding_Payment_Retention_Money=Current_Application_Temporary_Withholding_Payment_Retention_Money,
        Current_Application_Refund_of_Payment_Retention_Money=Current_Application_Refund_of_Payment_Retention_Money,
        Current_Application_Total_Amount_Receivable=Current_Application_Total_Amount_Receivable,
        Current_Application_Deduction_Penalty=Current_Application_Deduction_Penalty,
        Current_Application_Deduction_Fine=Current_Application_Deduction_Fine,
        Current_Application_Deduction_Other=Current_Application_Deduction_Other,
        Current_Application_Total_Deductions=Current_Application_Total_Deductions,
        Current_Application_Total_Accounts_Payable=Current_Application_Total_Accounts_Payable,
        # 本次审核参数
        Current_Approved_Internal_Contract_Project=Current_Approved_Internal_Contract_Project,  # 本期审定计量金额（从附件3 中提取）
        Current_Approved_Prepayment_Safety_Civilized_Construction_Fee=Current_Approved_Prepayment_Safety_Civilized_Construction_Fee,
        Current_Approved_Change_Project=Current_Approved_Change_Project,
        Current_Approved_Time_and_Materials_Project=Current_Approved_Time_and_Materials_Project,
        Current_Approved_Claim_Project=Current_Approved_Claim_Project,
        Current_Approved_Price_Adjustment=Current_Approved_Price_Adjustment,
        Current_Approved_Other=Current_Approved_Other,
        Current_Approved_Tax_Adjustment_Payable_Amount=Current_Approved_Tax_Adjustment_Payable_Amount,
        Current_Approved_Financial_Evaluation_Approval=Current_Approved_Financial_Evaluation_Approval,
        Current_Approved_Total_Completed_Output_Value=Current_Approved_Total_Completed_Output_Value,
        Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A=Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A,
        Current_Approved_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A=Current_Approved_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
        Current_Approved_Accounts_Payable_Advance_Payment=Current_Approved_Accounts_Payable_Advance_Payment,
        Current_Approved_Advance_Payment_Offset=Current_Approved_Advance_Payment_Offset,
        Current_Approved_Temporary_Withholding_Payment_Retention_Money=Current_Approved_Temporary_Withholding_Payment_Retention_Money,
        Current_Approved_Refund_of_Payment_Retention_Money=Current_Approved_Refund_of_Payment_Retention_Money,
        Current_Approved_Total_Amount_Receivable=Current_Approved_Total_Amount_Receivable,
        Current_Approved_Deduction_Penalty=Current_Approved_Deduction_Penalty,
        Current_Approved_Deduction_Fine=Current_Approved_Deduction_Fine,
        Current_Approved_Deduction_Other=Current_Approved_Deduction_Other,
        Current_Approved_Total_Deductions=Current_Approved_Total_Deductions,
        Current_Approved_Total_Accounts_Payable=Current_Approved_Total_Accounts_Payable
    )
    print(f"附件2已生成，完成时间为{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    fujian1_save_path = create_fujian1(
        output_path=output_folder_path,
        Contract_Name=Contract_Name,
        Contract_Number=project_info_dict['Contract_Number'],
        year=year, month=month,
        Reporting_Periods=Reporting_Periods,
        Construction_Unit=project_info_dict['Construction_Unit'],
        Contract_Amount=project_info_dict['Contract_Amount'],
        Funding_Source=project_info_dict['Funding_Source'],
        Consulting_Unit=project_info_dict['Consulting_Unit']
    )
    print(f"附件1已生成，完成时间为{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    doc = Document()
    doc.styles["Normal"].font.size = Pt(14)  # 设置字体大小为"四号"
    doc.styles["Normal"].font.name = "Times New Roman"  # 设置非中文字体为"新罗马"
    doc.styles["Normal"].element.rPr.rFonts.set(qn("w:eastAsia"), u"仿宋GB_2312")  # 设置中文字体为"仿宋GB_2312"
    report_file_name = jdk_xd(doc,
                              output_path=output_folder_path,
                              Contract_Name=Contract_Name,
                              Reporting_Periods=Reporting_Periods,
                              Consulting_Unit=project_info_dict['Consulting_Unit'],
                              Principal_Party=project_info_dict['Principal_Party'],
                              Design_Unit=project_info_dict['Design_Unit'],
                              Supervision_Unit=project_info_dict['Supervision_Unit'],
                              Construction_Unit=project_info_dict['Construction_Unit'],
                              Project_Overview=project_info_dict['Project_Overview'],
                              Price_Form=project_info_dict['Price_Form'],
                              Contract_Amount=project_info_dict['Contract_Amount'],
                              Current_Application_Internal_Contract_Project=Current_Application_Internal_Contract_Project,
                              Current_Approved_Internal_Contract_Project=Current_Approved_Internal_Contract_Project,
                              Current_Approved_Financial_Evaluation_Approval=Current_Approved_Financial_Evaluation_Approval,
                              Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A=Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A,
                              Total_Completed_Output_Value=Total_Completed_Output_Value,
                              Current_Approved_Total_Completed_Output_Value=Current_Application_Total_Completed_Output_Value,
                              Accounts_Payable_Advance_Payment=Accounts_Payable_Advance_Payment,
                              Advance_Payment_Offset=Advance_Payment_Offset,
                              Total_Accounts_Payable=Total_Accounts_Payable,
                              project_info_path=project_info_xlsx_dir,
                              )
    doc.save(report_file_name)
    # 保存附件3、4
    # print(fj3.filename, fj4.filename)

    fj_3_name = f'{Contract_Name}第{Reporting_Periods}期进度款单位工程审核汇总对比表（附件三）.xlsx'
    fj_4_name = f'{Contract_Name}第{Reporting_Periods}期进度款单位工程审核明细对比表（附件四）.xlsx'

    fj3_save_path = os.path.join(output_folder_path, fj_3_name)
    fj4_save_path = os.path.join(output_folder_path, fj_4_name)

    with open(fj3_save_path, 'wb') as f:
        f.write(fj3.file.read())
        f.close()
    with open(fj4_save_path, 'wb') as f:
        f.write(fj4.file.read())
        f.close()
    wb_fj4 = xl.load_workbook(fj4_save_path)
    ws_fj4 = wb_fj4.active
    # 设置表头
    ws_fj4['A2'] = f'单位工程审核明细对比表（第 {Reporting_Periods} 次进度付款）'
    ws_fj4[
        'A3'] = f'合同名称：{Contract_Name}' + " " * 30 + f"合同编号：{project_info_dict['Contract_Number']}" + " " * 30 + f"月份：{year}年第{month}月"
    ws_fj4['J4'] = f'本期申报（第{Reporting_Periods}次）'
    ws_fj4['L4'] = f'本期审核（第{Reporting_Periods}次）'
    wb_fj4.save(fj4_save_path)
    # print(fj3_save_path, fj4_save_path)
    end_time = time.time()
    print(f"已完成操作，整体耗时{round(end_time - start_time, 2)}")
    print(f'合同编号是：{project_info_dict["Contract_Number"]},期数是：{Reporting_Periods},年份是：{year},月份是：{month}')
    # 将output_folder_path文件夹下的文件打包成zip文件
    zip_file_name = f"{Contract_Name}_{Reporting_Periods}_{year}_{month}"
    zip_out_path = os.path.join(os.path.dirname(output_folder_path), zip_file_name)
    print(zip_out_path, output_folder_path)
    shutil.make_archive(zip_out_path, "zip", output_folder_path)
    # 将zip文件发送至前端
    return FileResponse(zip_out_path + ".zip", filename=zip_file_name + ".zip")


@app.get('/creat_reports_zip_not_first', response_class=HTMLResponse)
async def creat_reports_zip_not_first(request: Request):
    try:
        username = request.session.get('username')
    except:
        username = 'Guest'
    target_dir = os.path.join(os.path.dirname(__file__), "user_folder", f'{username}', "progress_reports_xd")
    contract_name = []
    for dir, folder, file in os.walk(target_dir):
        contract_name = folder
        break
    print(contract_name)
    return templates.TemplateResponse("Quick_progress_report_xd_not_first_time.html",
                                      {"request": request, "contract_name": contract_name})


@app.post('/creat_reports_xd_result_old_data', response_class=HTMLResponse)
async def creat_reports_xd_result_old_data(request: Request,
                                           Contract_Name=Form(default="测试"),
                                           Reporting_Periods=Form(default=1),
                                           Financial_Evaluation_Deduction_Last_Period: float = Form(default=0),
                                           file_input_ht: UploadFile = File(default=None),
                                           file_input_ss: UploadFile = File(default=None),
                                           file_input_sd: UploadFile = File(default=None)
                                           ):
    start_time = time.time()
    username = request.session.get('username')
    # 读取对应该项目的一些历史数据, 调用相关参数
    data_path = os.path.join(os.path.dirname(__file__), "function", "progress_reports_xd")
    project_name_dict = get_project_data_dict(output_path=data_path, Contract_Name=Contract_Name)
    Contract_Number = project_name_dict['Contract_Number']  # 合同编号
    Principal_Party = project_name_dict['Principal_Party']  # 建设单位
    Construction_Unit = project_name_dict['Construction_Unit']  # 施工单位
    Supervision_Unit = project_name_dict['Supervision_Unit']  # 监理单位
    Design_Unit = project_name_dict['Design_Unit']  # 设计单位
    Consulting_Unit = project_name_dict['Consulting_Unit']  # 咨询单位
    Project_Overview = project_name_dict['Project_Overview']  # 工程概况
    Funding_Source = project_name_dict['Funding_Source']  # 资金来源
    Price_Form = project_name_dict['Price_Form']  # 价格形式
    Progress_Payment_Ratio = project_name_dict['Progress_Payment_Ratio']  # 进度款支付比例
    Contract_Amount = project_name_dict['Contract_Amount']  # 合同金额
    Advance_Payment_Payment_Ratio = float(project_name_dict['Advance_Payment_Payment_Ratio']) / 100  # 预付款支付比例
    Advance_Payment_Deduction_Ratio = float(project_name_dict['Advance_Payment_Deduction_Ratio']) / 100  # 预付款扣除比例

    global project_dir_name, function_dir
    year = datetime.datetime.now().strftime("%Y")
    month = datetime.datetime.now().strftime("%m")
    # 创建一个用于保存主程序过程文件的文件夹，我将它命名为progress_reports_xd，意味对接新点软件的进度报告
    project_dir = os.path.join(os.path.dirname(__file__), "user_folder", f'{username}',
                               "progress_reports_xd", Contract_Name, f'第{Reporting_Periods}期')
    try:
        os.makedirs(project_dir)
    except FileExistsError:
        print(f'{project_dir}已存在')
        shutil.rmtree(project_dir)
        os.makedirs(project_dir, exist_ok=True)

    input_folder_path = creat_folder(os.path.join(project_dir, "input_files"))  # 输入文件路径
    output_folder_path = creat_folder(os.path.join(project_dir, "output_files"))  # 输出文件路径
    process_folder_path = creat_folder(os.path.join(project_dir, "process_files"))  # 过程文件路径

    ht_folder = creat_folder(os.path.join(input_folder_path, "ht"))  # 新建合同文件夹
    send_folder = creat_folder(os.path.join(input_folder_path, "ss"))  # 新建送审文件夹
    audited_folder = creat_folder(os.path.join(input_folder_path, "sd"))  # 新建审定文件夹

    # 解压合同文件至过程文件夹
    # 保存压缩文件至 input_files/合同 文件夹
    # print(file_input_ht.filename)
    ht_zip = Save_upload_file(file_input_ht, ht_folder, f"合同.zip")
    print(ht_zip)
    zip_unzip(zip_file_name=ht_zip, target_dir=ht_folder)
    send_zip = Save_upload_file(file_input_ss, send_folder, f"送审.zip")
    print(send_zip)
    zip_unzip(zip_file_name=send_zip, target_dir=send_folder)
    audited_zip = Save_upload_file(file_input_sd, audited_folder, f"审定.zip")
    print(audited_zip)
    zip_unzip(zip_file_name=audited_zip, target_dir=audited_folder)

    # 处理表格
    print(f'开始处理表格')
    get_file_path(ht_folder, process_folder_path, 'ht')
    get_file_path(send_folder, process_folder_path, 'ss')
    get_file_path(audited_folder, process_folder_path, 'sd')

    # 定义过程-ht拆分文件夹路径
    print('拆分文件夹')
    ht_excel_file_path_fb = creat_folder(os.path.join(process_folder_path, "ht", 'fb'))
    ht_excel_file_path_dw = creat_folder(os.path.join(process_folder_path, "ht", 'dw'))
    ht_excel_file_path_dj = creat_folder(os.path.join(process_folder_path, "ht", 'dj'))
    ht_excel_file_path_zj = creat_folder(os.path.join(process_folder_path, "ht", 'zj'))
    send_excel_file_path_fb = creat_folder(os.path.join(process_folder_path, "ss", 'fb'))
    send_excel_file_path_dw = creat_folder(os.path.join(process_folder_path, "ss", 'dw'))
    send_excel_file_path_dj = creat_folder(os.path.join(process_folder_path, "ss", 'dj'))
    send_excel_file_path_zj = creat_folder(os.path.join(process_folder_path, "ss", 'zj'))
    audited_excel_file_path_fb = creat_folder(os.path.join(process_folder_path, "sd", 'fb'))
    audited_excel_file_path_dw = creat_folder(os.path.join(process_folder_path, "sd", 'dw'))
    audited_excel_file_path_dj = creat_folder(os.path.join(process_folder_path, "sd", 'dj'))
    audited_excel_file_path_zj = creat_folder(os.path.join(process_folder_path, "sd", 'zj'))

    # 二次处理表格（删行）
    print('二次处理表格')
    for leixing in ["ht", "ss", "sd"]:
        leixing_path = os.path.join(process_folder_path, leixing)
        for file in os.listdir(leixing_path):
            excel_path = os.path.join(leixing_path, file)
            # print(excel_path)
            if file == "fb":
                fb_excel(excel_path)
            elif file == "dj":
                dj_excel(excel_path)
            elif file == "dw":
                dw_excel(excel_path)
            else:
                zj_excel(excel_path)
    Reporting_Periods = int(Reporting_Periods)
    print('开始处理连接上期数据')
    project_last_dir = os.path.join(os.path.dirname(__file__), "function", "progress_reports_xd", Contract_Name,
                                    f'第{Reporting_Periods - 1}期', "output_files")
    # 获取上一期的数据
    fj_3_name_last = f'{Contract_Name}第{Reporting_Periods - 1}期进度款单位工程审核汇总对比表（附件三）.xlsx'
    fj_4_name_last = f'{Contract_Name}第{Reporting_Periods - 1}期进度款单位工程审核明细对比表（附件四）.xlsx'
    fj_2_name_last = f'{Contract_Name}第{Reporting_Periods - 1}期进度款工程付款审核表（附件二）.xlsx'
    fj_1_name_last = f'{Contract_Name}第{Reporting_Periods - 1}期进度款付款汇总表（附件一）.xlsx'

    fj_1_path_last = os.path.join(project_last_dir, fj_1_name_last)
    fj_2_path_last = os.path.join(project_last_dir, fj_2_name_last)
    fj_3_path_last = os.path.join(project_last_dir, fj_3_name_last)
    fj_4_path_last = os.path.join(project_last_dir, fj_4_name_last)
    print('正在生成附件4')
    fujian4_save_path = fujian4(
        Contract_Name=Contract_Name, Reporting_Periods=Reporting_Periods, year=year, month=month,
        Contract_Number=Contract_Number,
        fenbu_path_ht=ht_excel_file_path_fb, djcs_path_ht=ht_excel_file_path_dj,
        zjcs_path_ht=ht_excel_file_path_zj, danwei_path_ht=ht_excel_file_path_dw,
        fenbu_path_sent=send_excel_file_path_fb, djcs_path_sent=send_excel_file_path_dj,
        zjcs_path_sent=send_excel_file_path_zj, danwei_path_sent=send_excel_file_path_dw,
        fenbu_path_audited=audited_excel_file_path_fb,
        djcs_path_audited=audited_excel_file_path_dj,
        zjcs_path_audited=audited_excel_file_path_zj,
        danwei_path_audited=audited_excel_file_path_dw,
        output_path=output_folder_path,
        input_ht_dx=ht_folder
    )

    print(f'fj_4_path_last:{fj_4_path_last}')
    wb_fujian4_last = xl.load_workbook(fj_4_path_last, data_only=True)
    ws_fujian4_last = wb_fujian4_last.active
    P_result = []
    Q_result = []
    for row in ws_fujian4_last.iter_rows():
        if row[0].row < 6:
            continue
        if row[column_index_from_string("P") - 1].value is None:
            P_result.append(0)
        else:
            p = round(float(row[column_index_from_string("P") - 1].value), 2)
            P_result.append(p)
        if row[column_index_from_string("Q") - 1].value is None:
            Q_result.append(0)
        else:
            q = round(float(row[column_index_from_string("Q") - 1].value), 2)
            Q_result.append(q)

    wb_fujian4 = xl.load_workbook(fujian4_save_path)
    ws_fujian4 = wb_fujian4.active

    for i, p in enumerate(P_result):
        ws_fujian4[f'H{i + 6}'] = p
    for i, q in enumerate(Q_result):
        ws_fujian4[f'I{i + 6}'] = q
    for row in ws_fujian4.iter_rows(min_row=6, max_row=ws_fujian4.max_row):
        gcl_sent_num = column_index_from_string("J") - 1
        gcl_audited_num = column_index_from_string("L") - 1
        price_sent_num = column_index_from_string("K") - 1
        price_audited_num = column_index_from_string("M") - 1

        gcl_sent = row[gcl_sent_num].value
        gcl_audited = row[gcl_audited_num].value
        price_sent = row[price_sent_num].value
        price_audited = row[price_audited_num].value

        write_col = row[0].row
        if gcl_sent is None or gcl_sent == '':
            gcl_sent = 0
        if gcl_audited is None or gcl_audited == '':
            gcl_audited = 0
        if price_sent is None or price_sent == '':
            price_sent = 0
        if price_audited is None or price_audited == '':
            price_audited = 0
        if gcl_audited - gcl_sent == 0:
            ws_fujian4[f'N{write_col}'] = ""
        else:
            ws_fujian4[f'N{write_col}'] = gcl_audited - gcl_sent
        if price_audited - price_sent == 0:
            ws_fujian4[f'O{write_col}'] = ""
        else:
            ws_fujian4[f'O{write_col}'] = price_audited - price_sent

    # 本期末累计完成
    for row in ws_fujian4.iter_rows(min_row=6, max_row=ws_fujian4.max_row):
        gcl_num = column_index_from_string("H") - 1
        price_num = column_index_from_string("I") - 1
        gcl_audited_num = column_index_from_string("L") - 1
        price_audited_num = column_index_from_string("M") - 1
        gcl = row[gcl_num].value
        price = row[price_num].value
        gcl_audited = row[gcl_audited_num].value
        price_audited = row[price_audited_num].value
        write_col = row[0].row
        if gcl is None or gcl == '':
            gcl = 0
        if price is None or price == '':
            price = 0
        if gcl_audited is None or gcl_audited == '':
            gcl_audited = 0
        if price_audited is None or price_audited == '':
            price_audited = 0
        if gcl + gcl_audited == 0:
            ws_fujian4[f'P{write_col}'] = ""
        else:
            ws_fujian4[f'P{write_col}'] = gcl + gcl_audited
        if price + price_audited == 0:
            ws_fujian4[f'Q{write_col}'] = ""
        else:
            ws_fujian4[f'Q{write_col}'] = price + price_audited

    wb_fujian4.save(fujian4_save_path)
    print(f"附件4已生成，完成时间为{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print('正在生成附件3')
    fujian3_save_path = create_fujian3(
        input_path=input_folder_path,
        output_path=output_folder_path,
        Contract_Name=Contract_Name,
        Contract_Number=Contract_Number,
        Reporting_Periods=Reporting_Periods,
        year=year,
        month=month
    )
    wb_fujian3_last = xl.load_workbook(fj_3_path_last)
    ws_fujian3_last = wb_fujian3_last.active
    # 读取附件3 的G列
    C_result = []
    G_result = []
    for row in ws_fujian3_last.iter_rows():
        if row[0].row < 6:
            continue
        C_result.append(row[column_index_from_string("C") - 1].value)
        G_result.append(row[column_index_from_string("G") - 1].value)
    wb_fujian3 = xl.load_workbook(fujian3_save_path)
    ws_fujian3 = wb_fujian3.active
    for i, g in enumerate(G_result):
        ws_fujian3[f'D{i + 6}'] = g
        ws_fujian3[f'G{i + 6}'] = ws_fujian3[f'F{i + 6}'].value + g
    for i, c in enumerate(C_result):
        ws_fujian3[f'C{i + 6}'] = c
    wb_fujian3.save(fujian3_save_path)
    print(f"附件3已生成，完成时间为{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    wb_fujian3 = xl.load_workbook(fujian3_save_path)
    ws_fujian3 = wb_fujian3.active
    Current_Application_Internal_Contract_Project = round(ws_fujian3[f'E{ws_fujian3.max_row}'].value, 2)
    Current_Approved_Internal_Contract_Project = round(ws_fujian3[f'F{ws_fujian3.max_row}'].value, 2)

    wb_fujian2_last = xl.load_workbook(fj_2_path_last, data_only=True)
    ws_fujian2_last = wb_fujian2_last.active
    Completed_Contract_Internal_Projects_Until_Last_Period = ws_fujian2_last['H6'].value
    Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period = ws_fujian2_last['H7'].value
    Accumulated_Completed_Amount_Until_Last_Period_Change_Projects = ws_fujian2_last['H8'].value
    Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project = ws_fujian2_last['H9'].value
    Accumulated_Completed_Amount_Until_Last_Period_Claim_Project = ws_fujian2_last['H10'].value
    Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment = ws_fujian2_last['H11'].value
    Accumulated_Completed_Amount_Until_Last_Period_Other = ws_fujian2_last['H12'].value
    Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount = ws_fujian2_last['H13'].value
    Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction = ws_fujian2_last["H14"].value
    Total_Completed_Output_Value = ws_fujian2_last['H15'].value
    Deduction_of_Supply_of_Plants_by_Party_A = ws_fujian2_last['H16'].value
    Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A = ws_fujian2_last['H17'].value
    Accounts_Payable_Advance_Payment = ws_fujian2_last['H18'].value
    Advance_Payment_Offset = ws_fujian2_last['H19'].value
    Temporary_Withholding_Payment_Retention_Money = ws_fujian2_last['H20'].value
    Refund_of_Payment_Retention_Money = ws_fujian2_last['H21'].value
    Total_Amount_Receivable = ws_fujian2_last['H22'].value
    Deduction_Penalty = ws_fujian2_last['H23'].value
    Deduction_Fine = ws_fujian2_last['H24'].value
    Deduction_Other = ws_fujian2_last['H25'].value
    Total_Deductions = ws_fujian2_last['H26'].value
    Total_Accounts_Payable = ws_fujian2_last['H27'].value

    # 计算预付款抵扣金额
    # 加个判断，如果本期累计支付金额超过
    Current_Approved_Total_Completed_Output_Value = float(Current_Approved_Internal_Contract_Project) - float(
        Financial_Evaluation_Deduction_Last_Period)
    price_leiji = round((float(Total_Completed_Output_Value) + float(Current_Approved_Total_Completed_Output_Value)), 2)
    leiji_amount = round((float(price_leiji) / float(Contract_Amount)) * 100, 2)
    if leiji_amount < 30:
        Current_Approved_Advance_Payment_Offset = 0
    elif 30 <= leiji_amount < 80:
        Current_Approved_Advance_Payment_Offset = round(
            (float(price_leiji) - float(Contract_Amount) * Advance_Payment_Payment_Ratio) / (float(Contract_Amount) * (
                    Advance_Payment_Deduction_Ratio - Advance_Payment_Payment_Ratio)) * float(
                Accounts_Payable_Advance_Payment) - float(
                Advance_Payment_Offset), 2)
    print(f'Current_Approved_Advance_Payment_Offset{Current_Approved_Advance_Payment_Offset}')
    html_response = templates.TemplateResponse("Quick_progress_report_xd_not_first_time_1_2.html",
                                               {
                                                   "request": request,
                                                   "Contract_Name": Contract_Name,
                                                   "Contract_Number": Contract_Number,
                                                   "Principal_Party": Principal_Party,
                                                   "Construction_Unit": Construction_Unit,
                                                   "Supervision_Unit": Supervision_Unit,
                                                   "Design_Unit": Design_Unit,
                                                   "Consulting_Unit": Consulting_Unit,
                                                   "Reporting_Periods": Reporting_Periods,
                                                   "Funding_Source": Funding_Source,
                                                   "Price_Form": Price_Form,
                                                   "Project_Overview": Project_Overview,
                                                   "Contract_Amount": Contract_Amount,
                                                   "Progress_Payment_Ratio": Progress_Payment_Ratio,
                                                   "Current_Application_Internal_Contract_Project": Current_Application_Internal_Contract_Project,
                                                   "Current_Application_Prepayment_Safety_Civilized_Construction_Fee": 0,
                                                   "Current_Application_Change_Project": 0,
                                                   "Current_Application_Time_and_Materials_Project": 0,
                                                   "Current_Application_Claim_Project": 0,
                                                   "Current_Application_Price_Adjustment": 0,
                                                   "Current_Application_Other": 0,
                                                   "Current_Application_Tax_Adjustment_Payable_Amount": 0,
                                                   "Current_Application_Financial_Evaluation_Reduction": 0,
                                                   "Current_Application_Total_Completed_Output_Value": 0,
                                                   "Current_Approved_Internal_Contract_Project": Current_Approved_Internal_Contract_Project,
                                                   "Current_Approved_Prepayment_Safety_Civilized_Construction_Fee": 0,
                                                   "Current_Approved_Change_Project": 0,
                                                   "Current_Approved_Time_and_Materials_Project": 0,
                                                   "Current_Approved_Claim_Project": 0,
                                                   "Current_Approved_Price_Adjustment": 0,
                                                   "Current_Approved_Other": 0,
                                                   "Current_Approved_Financial_Evaluation_Approval": Financial_Evaluation_Deduction_Last_Period,
                                                   "Current_Approved_Tax_Adjustment_Payable_Amount": 0,
                                                   "Current_Approved_Total_Completed_Output_Value": 0,
                                                   "year": year,
                                                   "month": month,
                                                   'Completed_Contract_Internal_Projects_Until_Last_Period': Completed_Contract_Internal_Projects_Until_Last_Period,
                                                   'Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period': Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period,
                                                   "Accumulated_Completed_Amount_Until_Last_Period_Change_Projects": Accumulated_Completed_Amount_Until_Last_Period_Change_Projects,
                                                   'Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project': Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project,
                                                   'Accumulated_Completed_Amount_Until_Last_Period_Claim_Project': Accumulated_Completed_Amount_Until_Last_Period_Claim_Project,
                                                   'Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment': Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment,
                                                   'Accumulated_Completed_Amount_Until_Last_Period_Other': Accumulated_Completed_Amount_Until_Last_Period_Other,
                                                   'Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount': Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount,
                                                   'Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction': Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction,
                                                   'Total_Completed_Output_Value': Total_Completed_Output_Value,
                                                   'Deduction_of_Supply_of_Plants_by_Party_A': Deduction_of_Supply_of_Plants_by_Party_A,
                                                   'Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A': Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
                                                   'Accounts_Payable_Advance_Payment': Accounts_Payable_Advance_Payment,
                                                   'Advance_Payment_Offset': Advance_Payment_Offset,
                                                   'Temporary_Withholding_Payment_Retention_Money': Temporary_Withholding_Payment_Retention_Money,
                                                   'Refund_of_Payment_Retention_Money': Refund_of_Payment_Retention_Money,
                                                   "Total_Amount_Receivable": Total_Amount_Receivable,
                                                   'Deduction_Penalty': Deduction_Penalty,
                                                   'Deduction_Fine': Deduction_Fine,
                                                   'Deduction_Other': Deduction_Other,
                                                   'Total_Deductions': Total_Deductions,
                                                   'Total_Accounts_Payable': Total_Accounts_Payable,
                                                   'Current_Approved_Advance_Payment_Offset': Current_Approved_Advance_Payment_Offset
                                               }
                                               )
    return html_response
    # return 'anquan'


@app.post("/creat_reports_xd_result_1_2_old", response_class=HTMLResponse)
async def creat_reports_xd_result(request: Request,
                                  Contract_Name=Form(...),
                                  Reporting_Periods=Form(...),
                                  # year=Form(...),
                                  # month=Form(...),
                                  Completed_Contract_Internal_Projects_Until_Last_Period: str = Form(default=0),
                                  Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period: str = Form(
                                      default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Change_Projects: str = Form(default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project: str = Form(
                                      default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Claim_Project: str = Form(default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment: str = Form(
                                      default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Other: str = Form(default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount: str = Form(
                                      default=0),
                                  Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction: str = Form(
                                      default=0),
                                  Total_Completed_Output_Value: str = Form(default=0),
                                  Deduction_of_Supply_of_Plants_by_Party_A: str = Form(default=0),
                                  Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A: str = Form(
                                      default=0),
                                  Accounts_Payable_Advance_Payment: str = Form(default=0),
                                  Advance_Payment_Offset: str = Form(default=0),
                                  Temporary_Withholding_Payment_Retention_Money: str = Form(default=0),
                                  Refund_of_Payment_Retention_Money: str = Form(default=0),
                                  Total_Amount_Receivable: str = Form(default=0),
                                  Deduction_Penalty: str = Form(default=0),
                                  Deduction_Fine: str = Form(default=0),
                                  Deduction_Other: str = Form(default=0),
                                  Total_Deductions: str = Form(default=0),
                                  Total_Accounts_Payable: str = Form(default=0),
                                  # 本次送审的参数
                                  Current_Application_Internal_Contract_Project: str = Form(default=0),  #
                                  Current_Application_Prepayment_Safety_Civilized_Construction_Fee: str = Form(
                                      default=0),
                                  Current_Application_Change_Project: str = Form(default=0),
                                  Current_Application_Time_and_Materials_Project: str = Form(default=0),
                                  Current_Application_Claim_Project: str = Form(default=0),
                                  Current_Application_Price_Adjustment: str = Form(default=0),
                                  Current_Application_Other: str = Form(default=0),
                                  Current_Application_Tax_Adjustment_Payable_Amount: str = Form(default=0),
                                  Current_Application_Financial_Evaluation_Reduction: str = Form(default=0),
                                  Current_Application_Total_Completed_Output_Value: str = Form(default=0),
                                  Current_Application_Deduction_of_Supply_of_Plants_by_Party_A: str = Form(default=0),
                                  Current_Application_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A: str = Form(
                                      default=0),
                                  Current_Application_Accounts_Payable_Advance_Payment: str = Form(default=0),
                                  Current_Application_Advance_Payment_Offset: str = Form(default=0),
                                  Current_Application_Temporary_Withholding_Payment_Retention_Money: str = Form(
                                      default=0),
                                  Current_Application_Refund_of_Payment_Retention_Money: str = Form(default=0),
                                  Current_Application_Total_Amount_Receivable: str = Form(default=0),
                                  Current_Application_Deduction_Penalty: str = Form(default=0),
                                  Current_Application_Deduction_Fine: str = Form(default=0),
                                  Current_Application_Deduction_Other: str = Form(default=0),
                                  Current_Application_Total_Deductions: str = Form(default=0),
                                  Current_Application_Total_Accounts_Payable: str = Form(default=0),
                                  # 本次审核参数
                                  Current_Approved_Internal_Contract_Project: str = Form(default=0),
                                  Current_Approved_Prepayment_Safety_Civilized_Construction_Fee: str = Form(default=0),
                                  Current_Approved_Change_Project: str = Form(default=0),
                                  Current_Approved_Time_and_Materials_Project: str = Form(default=0),
                                  Current_Approved_Claim_Project: str = Form(default=0),
                                  Current_Approved_Price_Adjustment: str = Form(default=0),
                                  Current_Approved_Other: str = Form(default=0),
                                  Current_Approved_Tax_Adjustment_Payable_Amount: str = Form(default=0),
                                  Current_Approved_Financial_Evaluation_Approval: str = Form(default=0),
                                  Current_Approved_Total_Completed_Output_Value: str = Form(default=0),
                                  Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A: str = Form(default=0),
                                  Current_Approved_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A: str = Form(
                                      default=0),
                                  Current_Approved_Accounts_Payable_Advance_Payment: str = Form(default=0),
                                  Current_Approved_Advance_Payment_Offset: str = Form(default=0),
                                  Current_Approved_Temporary_Withholding_Payment_Retention_Money: str = Form(default=0),
                                  Current_Approved_Refund_of_Payment_Retention_Money: str = Form(default=0),
                                  Current_Approved_Total_Amount_Receivable: str = Form(default=0),
                                  Current_Approved_Deduction_Penalty: str = Form(default=0),
                                  Current_Approved_Deduction_Fine: str = Form(default=0),
                                  Current_Approved_Deduction_Other: str = Form(default=0),
                                  Current_Approved_Total_Deductions: str = Form(default=0),
                                  Current_Approved_Total_Accounts_Payable: str = Form(default=0)):
    start_time = time.time()
    year = datetime.datetime.now().strftime("%Y")
    month = datetime.datetime.now().strftime("%m")
    print(year, month, Contract_Name, Reporting_Periods)
    project_dir = os.path.join(os.path.dirname(__file__), "function", "progress_reports_xd", Contract_Name,
                               f'第{Reporting_Periods}期')
    project_info_xlsx_dir = os.path.join(os.path.dirname(__file__), "function", "progress_reports_xd")
    output_folder_path = creat_folder(os.path.join(project_dir, "output_files"))  # 输出文件路径

    project_info_dict = get_project_data_dict(output_path=project_info_xlsx_dir, Contract_Name=Contract_Name)

    fujian2_save_path = create_fujian2(
        output_path=output_folder_path,
        Contract_Name=Contract_Name, Contract_Number=project_info_dict['Contract_Number'],
        Contract_Amount=project_info_dict['Contract_Amount'], year=year, month=month,
        Reporting_Periods=Reporting_Periods,
        Advance_Payment_Payment_Ratio=project_info_dict['Advance_Payment_Payment_Ratio'],
        Advance_Payment_Deduction_Ratio=project_info_dict['Advance_Payment_Deduction_Ratio'],
        Completed_Contract_Internal_Projects_Until_Last_Period=Completed_Contract_Internal_Projects_Until_Last_Period,
        Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period=Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period,
        Accumulated_Completed_Amount_Until_Last_Period_Change_Projects=Accumulated_Completed_Amount_Until_Last_Period_Change_Projects,
        Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project=Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project,
        # 计日工项目
        Accumulated_Completed_Amount_Until_Last_Period_Claim_Project=Accumulated_Completed_Amount_Until_Last_Period_Claim_Project,
        # 索赔项目
        Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment=Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment,
        Accumulated_Completed_Amount_Until_Last_Period_Other=Accumulated_Completed_Amount_Until_Last_Period_Other,
        Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount=Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount,
        Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction=Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction,
        Total_Completed_Output_Value=Total_Completed_Output_Value,
        Deduction_of_Supply_of_Plants_by_Party_A=Deduction_of_Supply_of_Plants_by_Party_A,
        Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A=Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
        Accounts_Payable_Advance_Payment=Accounts_Payable_Advance_Payment,
        Advance_Payment_Offset=Advance_Payment_Offset,
        Temporary_Withholding_Payment_Retention_Money=Temporary_Withholding_Payment_Retention_Money,
        Refund_of_Payment_Retention_Money=Refund_of_Payment_Retention_Money,  # 返还支付保留金
        Total_Amount_Receivable=Total_Amount_Receivable,
        Deduction_Penalty=Deduction_Penalty,
        Deduction_Fine=Deduction_Fine,
        Deduction_Other=Deduction_Other,
        Total_Deductions=Total_Deductions,
        Total_Accounts_Payable=Total_Accounts_Payable,
        Current_Application_Internal_Contract_Project=Current_Application_Internal_Contract_Project,
        Current_Application_Prepayment_Safety_Civilized_Construction_Fee=Current_Application_Prepayment_Safety_Civilized_Construction_Fee,
        # 预付安全文明施工费
        Current_Application_Change_Project=Current_Application_Change_Project,  # 变更项目
        Current_Application_Time_and_Materials_Project=Current_Application_Time_and_Materials_Project,  # 计日工项目
        Current_Application_Claim_Project=Current_Application_Claim_Project,  # 索赔项目
        Current_Application_Price_Adjustment=Current_Application_Price_Adjustment,
        Current_Application_Other=Current_Application_Other,
        Current_Application_Tax_Adjustment_Payable_Amount=Current_Application_Tax_Adjustment_Payable_Amount,
        Current_Application_Financial_Evaluation_Reduction=Current_Application_Financial_Evaluation_Reduction,
        Current_Application_Total_Completed_Output_Value=Current_Application_Total_Completed_Output_Value,
        Current_Application_Deduction_of_Supply_of_Plants_by_Party_A=Current_Application_Deduction_of_Supply_of_Plants_by_Party_A,
        Current_Application_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A=Current_Application_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
        Current_Application_Accounts_Payable_Advance_Payment=Current_Application_Accounts_Payable_Advance_Payment,
        Current_Application_Advance_Payment_Offset=Current_Application_Advance_Payment_Offset,
        Current_Application_Temporary_Withholding_Payment_Retention_Money=Current_Application_Temporary_Withholding_Payment_Retention_Money,
        Current_Application_Refund_of_Payment_Retention_Money=Current_Application_Refund_of_Payment_Retention_Money,
        Current_Application_Total_Amount_Receivable=Current_Application_Total_Amount_Receivable,
        Current_Application_Deduction_Penalty=Current_Application_Deduction_Penalty,
        Current_Application_Deduction_Fine=Current_Application_Deduction_Fine,
        Current_Application_Deduction_Other=Current_Application_Deduction_Other,
        Current_Application_Total_Deductions=Current_Application_Total_Deductions,
        Current_Application_Total_Accounts_Payable=Current_Application_Total_Accounts_Payable,
        # 本次审核参数
        Current_Approved_Internal_Contract_Project=Current_Approved_Internal_Contract_Project,  # 本期审定计量金额（从附件3 中提取）
        Current_Approved_Prepayment_Safety_Civilized_Construction_Fee=Current_Approved_Prepayment_Safety_Civilized_Construction_Fee,
        Current_Approved_Change_Project=Current_Approved_Change_Project,
        Current_Approved_Time_and_Materials_Project=Current_Approved_Time_and_Materials_Project,
        Current_Approved_Claim_Project=Current_Approved_Claim_Project,
        Current_Approved_Price_Adjustment=Current_Approved_Price_Adjustment,
        Current_Approved_Other=Current_Approved_Other,
        Current_Approved_Tax_Adjustment_Payable_Amount=Current_Approved_Tax_Adjustment_Payable_Amount,
        Current_Approved_Financial_Evaluation_Approval=Current_Approved_Financial_Evaluation_Approval,
        Current_Approved_Total_Completed_Output_Value=Current_Approved_Total_Completed_Output_Value,
        Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A=Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A,
        Current_Approved_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A=Current_Approved_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
        Current_Approved_Accounts_Payable_Advance_Payment=Current_Approved_Accounts_Payable_Advance_Payment,
        Current_Approved_Advance_Payment_Offset=Current_Approved_Advance_Payment_Offset,
        Current_Approved_Temporary_Withholding_Payment_Retention_Money=Current_Approved_Temporary_Withholding_Payment_Retention_Money,
        Current_Approved_Refund_of_Payment_Retention_Money=Current_Approved_Refund_of_Payment_Retention_Money,
        Current_Approved_Total_Amount_Receivable=Current_Approved_Total_Amount_Receivable,
        Current_Approved_Deduction_Penalty=Current_Approved_Deduction_Penalty,
        Current_Approved_Deduction_Fine=Current_Approved_Deduction_Fine,
        Current_Approved_Deduction_Other=Current_Approved_Deduction_Other,
        Current_Approved_Total_Deductions=Current_Approved_Total_Deductions,
        Current_Approved_Total_Accounts_Payable=Current_Approved_Total_Accounts_Payable
    )
    print(f"附件2已生成，完成时间为{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    fujian1_save_path = create_fujian1(
        output_path=output_folder_path,
        Contract_Name=Contract_Name,
        Contract_Number=project_info_dict['Contract_Number'],
        year=year, month=month,
        Reporting_Periods=Reporting_Periods,
        Construction_Unit=project_info_dict['Construction_Unit'],
        Contract_Amount=project_info_dict['Contract_Amount'],
        Funding_Source=project_info_dict['Funding_Source'],
        Consulting_Unit=project_info_dict['Consulting_Unit']
    )
    print(f"附件1已生成，完成时间为{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    doc = Document()
    doc.styles["Normal"].font.size = Pt(14)  # 设置字体大小为"四号"
    doc.styles["Normal"].font.name = "Times New Roman"  # 设置非中文字体为"新罗马"
    doc.styles["Normal"].element.rPr.rFonts.set(qn("w:eastAsia"), u"仿宋GB_2312")  # 设置中文字体为"仿宋GB_2312"
    report_file_name = jdk_xd(doc,
                              output_path=output_folder_path,
                              Contract_Name=Contract_Name,
                              Reporting_Periods=Reporting_Periods,
                              Consulting_Unit=project_info_dict['Consulting_Unit'],
                              Principal_Party=project_info_dict['Principal_Party'],
                              Design_Unit=project_info_dict['Design_Unit'],
                              Supervision_Unit=project_info_dict['Supervision_Unit'],
                              Construction_Unit=project_info_dict['Construction_Unit'],
                              Project_Overview=project_info_dict['Project_Overview'],
                              Price_Form=project_info_dict['Price_Form'],
                              Contract_Amount=project_info_dict['Contract_Amount'],
                              Current_Application_Internal_Contract_Project=Current_Application_Internal_Contract_Project,
                              Current_Approved_Internal_Contract_Project=Current_Approved_Internal_Contract_Project,
                              Current_Approved_Financial_Evaluation_Approval=Current_Approved_Financial_Evaluation_Approval,
                              Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A=Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A,
                              Total_Completed_Output_Value=Total_Completed_Output_Value,
                              Current_Approved_Total_Completed_Output_Value=Current_Application_Total_Completed_Output_Value,
                              Accounts_Payable_Advance_Payment=Accounts_Payable_Advance_Payment,
                              Advance_Payment_Offset=Advance_Payment_Offset,
                              Total_Accounts_Payable=Total_Accounts_Payable,
                              project_info_path=project_info_xlsx_dir
                              )
    doc.save(report_file_name)
    end_time = time.time()
    print(f"已完成操作，整体耗时{round(end_time - start_time, 2)}")
    print(f'合同编号是：{project_info_dict["Contract_Number"]},期数是：{Reporting_Periods},年份是：{year},月份是：{month}')
    # 将output_folder_path文件夹下的文件打包成zip文件
    zip_file_name = f"{Contract_Name}_{Reporting_Periods}_{year}_{month}"
    zip_out_path = os.path.join(os.path.dirname(output_folder_path), zip_file_name)
    print(zip_out_path, output_folder_path)
    shutil.make_archive(zip_out_path, "zip", output_folder_path)
    # 将zip文件发送至前端
    return FileResponse(zip_out_path + ".zip", filename=zip_file_name + ".zip")

if __name__ == '__main__':
    import uvicorn

    uvicorn.run(app='main:app', host="0.0.0.0", port=8000, reload=True)
