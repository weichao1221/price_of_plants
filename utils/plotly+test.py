import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
import openpyxl as xl
import os
from openpyxl.utils import get_column_letter, column_index_from_string

excel_dir = os.path.dirname(__file__)
excel_name = '苗木.xlsx'

wb = xl.load_workbook(os.path.join(excel_dir, excel_name), data_only=True)
ws = wb['测试苗木']
start_name = "白皮松"  # 开始的名字
column_data = []
company_list = []
x_label = []
for column in range(column_index_from_string("J"), ws.max_column + 1):
    for row in range(2, 3):
        company_list.append(ws.cell(row=row, column=column).value)
    column_values = []
    for row in range(4, ws.max_row + 1):
        this_name = ws.cell(row=row, column=2).value
        if this_name != start_name:
            continue
        cell_value = ws.cell(row=row, column=column).value
        if cell_value is None:
            cell_value = 0
        column_values.append(cell_value)
        x_label_value = ws.cell(row=row, column=column_index_from_string("D")).value
        x_label.append(x_label_value)
    column_data.append(column_values)

# 创建一个 Plotly 图表
fig = go.Figure()

# 循环处理每个公司的数据
for index, column_values in enumerate(column_data, start=1):
    x = []
    for i, y in enumerate(column_values):
        x.append(i + 1)

    company_name = company_list[index - 1]  # 注意索引从0开始

    def exponential_func(x, a, b):
        return a * np.exp(b * x)
    print(x, column_values)
    x = np.array(x, dtype=float)
    y = np.array(column_values, dtype=float)

    # 筛选非零值
    x_filtered = x[y != 0]
    y_filtered = y[y != 0]

    if x_filtered.size == 0:
        continue

    # 进行曲线拟合
    popt, pcov = curve_fit(exponential_func, x_filtered, y_filtered)
    a, b = popt

    # 计算拟合曲线的R平方值
    y_predict = exponential_func(x_filtered, a, b)
    r_square = round(r2_score(y_filtered, y_predict), 5)

    # 添加散点图
    fig.add_trace(go.Scatter(x=x_filtered, y=y_filtered, mode='markers', name=f'{company_name}'))

    # 添加拟合曲线
    fig.add_trace(go.Scatter(x=x_filtered, y=y_predict, mode='lines', name=f'{company_name} (R^2={r_square:.2f})'))

# 设置图表布局
fig.update_layout(
    title=f'{start_name} 数据拟合',
    xaxis_title='X',
    yaxis_title='Y'
)

# 保存图表为 HTML 文件
fig.write_html("fig_html.html")
# 显示图表
fig.show()
