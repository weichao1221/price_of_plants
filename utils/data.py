import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import matplotlib.pyplot as plt
import numpy as np
from scipy.optimize import curve_fit
from scipy.stats import linregress

excel_dir = os.path.dirname(__file__)
excel_name = '苗木.xlsx'

wb = xl.load_workbook(os.path.join(excel_dir, excel_name), data_only=True)
ws = wb['测试苗木']



# 获取不考虑公司曲线的基础数据


def get_data_all(start_name):
    x_data = []
    y_data = []
    x_label = []
    count = 0
    for row in ws.iter_rows():
        if row[0].row in range(1, 4):
            continue
        # 判断苗木名称
        this_name = row[1].value
        if this_name != start_name: # 如果不是开始的名字，跳过
            continue
        count += 1
        x_data.append(count)
        x_label.append(row[3].value)
        y_value = [cell.value if cell.value is not None else 0 for cell in row[9:]]
        y_data.append(y_value)
    x_list = []
    y_list = []
    for i, y_value in enumerate(y_data, start=1):
        x_values = np.array([i] * len(y_value))  # 将 x 数据构造成一维数组
        y_values = np.array(y_value)  # 将 y 数据构造成一维数组
        for i in x_values:
            x_list.append(i)
        for j in y_values:
            y_list.append(j)

    x_list = np.array(x_list)
    y_list = np.array(y_list)

    x_filtered = x_list[y_list != 0]  # 跳过0值之后的x_data
    y_filtered = y_list[y_list != 0]  # 跳过0值之后对应的y_data

    return x_filtered, y_filtered

# 获取各公司曲线的基础数据


def get_data_company(start_name):
    column_data = []
    xlist = []
    ylist = []
    for column in range(column_index_from_string("J"), ws.max_column + 1):
        column_values = []
        for row in range(4, ws.max_row + 1):
            this_name = ws.cell(row=row, column=2).value
            if this_name != start_name:
                continue
            cell_value = ws.cell(row=row, column=column).value
            if cell_value is None:
                cell_value = 0
            column_values.append(cell_value)
        column_data.append(column_values)
    for index, column_values in enumerate(column_data, start=1):
        x = []
        for i, y in enumerate(column_values):
            x.append(i + 1)
        x = np.array(x)
        y = np.array(column_values)
        xlist.append(x)
        ylist.append(y)

    return xlist, ylist

# 获取公司序号和名称的字典


def company_name_dict():
    company_list = []
    for column in range(column_index_from_string("J"), ws.max_column + 1):
        for row in range(2, 3):
            company_list.append(ws.cell(row=row, column=column).value)
    company_dict = {}
    for i, name in enumerate(company_list):
        company_dict[i + 1] = name
    return company_dict


# 获取规格序号和名称的字典


def x_label_dict(x_label_list):
    x_label_dict = {}
    for i, label in enumerate(x_label_list):
        x_label_dict[i + 1] = label
    return x_label_dict


# 获取x轴标签  苗木规格数据
def get_x_label(start_name):
    x_label = []
    for row in ws.iter_rows():
        if row[0].row in range(1, 4):
            continue
        this_name = row[1].value
        if this_name != start_name:
            continue
        x_label.append(row[3].value)
    # x_label_dict = {}
    # for i, label in enumerate(x_label):
    #     x_label_dict[i + 1] = label
    return x_label


def get_x_data(start_name):
    x_data = []
    count = 0
    for row in ws.iter_rows():
        if row[0].row in range(1, 4):
            continue
        this_name = row[1].value
        if this_name != start_name:
            continue
        count += 1
        x_data.append(count)
    return x_data


# 定义指数函数模型
def exponential_model(x, a, b, c):
    return a * np.exp(b * x) + c

# 定义2阶多项式函数模型
def polynomial_model(x, a, b, c):
    return a * x**2 + b * x + c

# 获取指数拟合的R平方值
def get_exponential_r_squared(x_data, y_data):
    try:
        params, covariance = curve_fit(exponential_model, x_data, y_data, maxfev=2000)
        y_fit = exponential_model(x_data, *params)
        ss_res = np.sum((y_data - y_fit) ** 2)
        ss_tot = np.sum((y_data - np.mean(y_data)) ** 2)
        r_squared = 1 - (ss_res / ss_tot)
        return round(r_squared, 5)
    except RuntimeError:
        return 0


def get_polynomial_r_squared(x_data, y_data):
    params = np.polyfit(x_data, y_data, 2)
    y_fit = np.polyval(params, x_data)
    slope, intercept, r_value, p_value, std_err = linregress(y_fit, y_data)
    return round(r_value ** 2, 5)

def get_name_list():
    start_name_list = []
    for row in ws.iter_rows():
        if row[0].row in range(1, 4):
            continue
        if not row[1].value:
            continue
        start_name = row[1].value
        start_name_list.append(start_name)

    # start_name_list 去重
    start_name_list = list(set(start_name_list))
    # 对列表内容按照英文首字母顺序排序
    start_name_list.sort(key=lambda x: x[0])
    return start_name_list


def draw(start_name):
    x_label = get_x_label(start_name)
    x_data_index = get_x_data(start_name)
    xlist = get_data_company(start_name)[0]
    ylist = get_data_company(start_name)[1]
    marge_list = zip(xlist, ylist)
    plt.figure(figsize=(25.6, 14.4), dpi=200)
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    for index, (x, y) in enumerate(marge_list, start=1):
        # print(f"第{index}个已经开始绘制。。。")

        xvalue = x[y != 0]
        yvalue = y[y != 0]

        if xvalue.size == 0:
            continue
        exp_r_squared = get_exponential_r_squared(xvalue, yvalue)
        poly_r_squared = get_polynomial_r_squared(xvalue, yvalue)
        # print("指数拟合的R平方值:", exp_r_squared)
        # print("2阶多项式拟合的R平方值:", poly_r_squared)
        if exp_r_squared >= poly_r_squared:
            params, covariance = curve_fit(exponential_model, xvalue, yvalue, maxfev=2000)
            x_fit = np.linspace(min(xvalue), max(xvalue), 100)
            y_fit = exponential_model(x_fit, *params)
            a, b, c = params
            equation = f'曲线公式: y = {round(a, 2)} * e^{round(b, 2)}x + {round(c, 2)}'
            plt.scatter(xvalue, yvalue, label=f'{company_name_dict()[index]}')
            plt.plot(x_fit, y_fit, label=f'{company_name_dict()[index]}，{equation}，R^2={exp_r_squared}')
            plt.legend()
            # print(f'{start_name}-{company_name_dict()[index]}已完成绘制')
        else:
            params, covariance = curve_fit(polynomial_model, xvalue, yvalue, maxfev=2000)
            x_fit = np.linspace(min(xvalue), max(xvalue), 100)
            y_fit = polynomial_model(x_fit, *params)
            a, b, c = params
            equation = f'曲线公式: y = {a:.2f}x^2 + {b:.2f}x + {c:.2f}'
            plt.scatter(xvalue, yvalue, label=f'{company_name_dict()[index]}')
            plt.plot(x_fit, y_fit, label=f'{company_name_dict()[index]}，{equation}，R^2={poly_r_squared}')
            plt.legend()
            # print(f'{start_name}-{company_name_dict()[index]}已完成绘制')

    x_data = get_data_all(start_name)[0]
    y_data = get_data_all(start_name)[1]
    exp_r_squared = get_exponential_r_squared(x_data, y_data)
    poly_r_squared = get_polynomial_r_squared(x_data, y_data)
    # print("指数拟合的R平方值:", exp_r_squared)
    # print("2阶多项式拟合的R平方值:", poly_r_squared)
    if exp_r_squared >= poly_r_squared:
        params, covariance = curve_fit(exponential_model, x_data, y_data, maxfev=2000)
        x_fit = np.linspace(min(x_data), max(x_data), 100)
        y_fit = exponential_model(x_fit, *params)
        a, b, c = params
        equation = f'曲线公式: y = {round(a, 2)} * e^{round(b, 2)}x + {round(c, 2)}'
        # plt.scatter(x_data, y_data, label=f'综合考虑全部数据')
        plt.plot(x_fit, y_fit, label=f'综合考虑全部数据，{equation}，R^2={exp_r_squared}',linewidth=5, color="red")
        plt.legend()
    else:
        params, covariance = curve_fit(polynomial_model, x_data, y_data, maxfev=2000)
        x_fit = np.linspace(min(x_data), max(x_data), 100)
        y_fit = polynomial_model(x_fit, *params)
        a, b, c = params
        equation = f'曲线公式: y = {a:.2f}x^2 + {b:.2f}x + {c:.2f}'
        # plt.scatter(x_data, y_data, label=f'综合考虑全部数据')
        plt.plot(x_fit, y_fit, label=f'综合考虑全部数据，{equation}，R^2={poly_r_squared}',linewidth=5, color="red")
        plt.legend()

    plt.xlabel('规格')
    plt.ylabel('价格')
    plt.xticks(x_data_index, x_label)
    plt.savefig(f'main/{start_name}.png')
    # plt.show()
    # print("已完成画图")


if __name__ == '__main__':
    start_name = "白皮松"
    draw(start_name)
