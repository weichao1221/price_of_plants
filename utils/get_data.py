import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
import openpyxl as xl
import os
import time
from openpyxl.utils import get_column_letter, column_index_from_string

excel_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources')
excel_name = '苗木信息价整合.xlsx'

wb = xl.load_workbook(os.path.join(excel_dir, excel_name), data_only=True)
ws = wb.active


def get_name_list():
    name_list = []
    for row in ws.iter_rows():
        if row[0].row == 1:
            continue
        if row[1].value is None:
            continue
        name_list.append(row[1].value)
    new_list = list(set(name_list))
    new_list.sort(key=name_list.index)
    usefull_list = []
    for name in new_list:
        try:
            draw(get_data(name)[0], get_data(name)[1], name)
            usefull_list.append(name)
        # 如果出现任何错误，就跳过
        except Exception as e:
            continue

    # 返回四个列表
    changlv_name_list = []
    luoye_name_list = []
    guanmu_name_list = []
    dibeilei_name_list = []
    name_dict = get_name_dict()
    for name in usefull_list:
        if name_dict[name] == "常绿乔木":
            changlv_name_list.append(name)
        if name_dict[name] == "落叶乔木":
            luoye_name_list.append(name)
        if name_dict[name] == "落叶灌木":
            guanmu_name_list.append(name)
        if name_dict[name] == "地被类":
            dibeilei_name_list.append(name)
    return usefull_list, changlv_name_list, luoye_name_list, guanmu_name_list, dibeilei_name_list, get_name_dict()


def get_name_dict():
    name_dict = {}
    for row in ws.iter_rows():
        if row[0].row == 1:
            continue
        if row[1].value is None:
            continue
        name_dict[row[1].value] = row[6].value
    return name_dict


def get_x_label(start_name):
    x_label = []
    for row in ws.iter_rows():
        if row[0].row == 1:
            continue
        if row[1].value is None:
            continue
        if row[1].value != start_name:
            continue
        # column_data.append(row[4].value)
        x_label.append(row[2].value)
    return x_label


def get_data(start_name):
    column_data = []
    x_label = []
    for row in ws.iter_rows():
        if row[0].row == 1:
            continue
        if row[1].value is None:
            continue
        if row[1].value != start_name:
            continue
        column_data.append(row[4].value)
        x_label.append(row[2].value)
    x = list()
    y = list()
    for index, values in enumerate(column_data, start=1):
        x.append(index)
        y.append(values)
    return x, y


def draw_plotly(start_name):
    column_data = []
    x_label = []
    for row in ws.iter_rows():
        if row[0].row == 1:
            continue
        if row[1].value is None:
            continue
        if row[1].value != start_name:
            continue
        column_data.append(row[4].value)
        x_label.append(row[2].value)
    # print(x_label)

    # 创建一个 Plotly 图表
    fig = go.Figure()

    # 循环处理每个公司的数据
    x = list()
    y = list()
    for index, values in enumerate(column_data, start=1):
        # print(index, values)
        x.append(index)
        y.append(values)

    def exponential_model(x, a, b, c):
        return a * np.exp(b * x) + c

    x = np.array(x, dtype=float)
    y = np.array(y, dtype=float)

    # 筛选非零值
    x_data = x[y != 0]
    y_data = y[y != 0]

    params, covariance = curve_fit(exponential_model, x_data, y_data, maxfev=5000)
    y_predict = exponential_model(x_data, *params)
    ss_res = np.sum((y_data - y_predict) ** 2)
    ss_tot = np.sum((y_data - np.mean(y_data)) ** 2)
    a, b, c = params
    r_squared = round(1 - (ss_res / ss_tot), 3)
    equation = f'曲线公式: y = {round(a, 2)} * e^{round(b, 2)}x + {round(c, 2)}'
    # 添加散点图
    fig.add_trace(go.Scatter(x=x_data, y=y_data, mode='markers', name=f'原始散点图'))
    # 添加拟合曲线
    fig.add_trace(go.Scatter(x=x_data, y=y_predict, mode='lines', name=f'{equation} (R^2={r_squared:.2f})'))
    # 设置图表布局
    fig.update_layout(
        # width=800,
        # height=600,
        title=f'{start_name} 数据拟合曲线',
        xaxis_title='苗木规格',
        yaxis_title='苗木单价',
        xaxis=dict(
            tickmode='array',
            tickvals=x_data,
            ticktext=x_label,
        ),
    )
    return fig


def exponential_function(x, a, b):
    return a * np.exp(b * x)


def draw(x, y, start_name):
    # 拟合数据
    params, covariance = curve_fit(exponential_function, x, y)

    # 获取拟合后的参数
    a_fit, b_fit = params

    # 获取协方差矩阵对角元素，这些元素代表参数的方差
    variance_a, variance_b = np.diag(covariance)

    # 创建用于绘图的 x 值
    x_fit = np.linspace(min(x), max(x), 100)

    # 计算拟合后的 y 值
    y_fit = exponential_function(x_fit, a_fit, b_fit)

    # 创建散点图
    scatter = go.Scatter(x=x, y=y, mode='markers', name='散点')

    nihe = f'y = {a_fit:.2f} * e^({b_fit:.2f} * x), Var(a) = {variance_a:.2f}, Var(b) = {variance_b:.2f}'
    # 创建拟合曲线
    fit_curve = go.Scatter(x=x_fit, y=y_fit, mode='lines',
                           name=f'拟合曲线')

    x_label = get_x_label(start_name)
    # 创建图布局
    layout = go.Layout(
        title=f'{start_name}拟合曲线图',
        yaxis=dict(title='苗木价格'),
        xaxis=dict(
            title='苗木规格',
            tickmode='array',
            tickvals=x,
            ticktext=x_label,
        ))

    # 绘制图形
    fig = go.Figure(data=[scatter, fit_curve], layout=layout)
    return fig


if __name__ == '__main__':
    print(get_name_list()[3])
