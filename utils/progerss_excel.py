# # -*- coding: utf-8 -*-
import openpyxl as pl
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import xlrd
from utils.Excel_usual import convert_xls_to_xlsx
from utils.excels.num_to_cn import number_to_chinese

# 拆分分部分项表格
def split_excel_fenbu(input_file, process_path, category, part):
    wb = pl.load_workbook(input_file, data_only=True)
    for index, sheet_name in enumerate(wb.sheetnames):
        # 创建一个新的工作簿
        output_wb = pl.Workbook()
        # 获取当前工作表
        source_sheet = wb[sheet_name]
        # 在新工作簿中创建相应的工作表
        output_sheet = output_wb.active
        ws = wb[sheet_name]
        output_sheet.title = part  # 以后可以改成sheet_name
        # 复制数据到新的工作表
        for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
            output_sheet.append(row)
        # 在表的第一列前拆入一列
        output_sheet.insert_cols(1)
        # 将这一列的中单元格数据都填充为ws['A2'].value
        for i in range(2, output_sheet.max_row + 1):
            # 其中的输入都包含"工程名称："，去掉这个字符串
            output_sheet[f'A{i}'] = ws['A2'].value.replace('工程名称：', '')
        # 删除E列和I列
        # output_sheet.delete_cols(column_index_from_string("H"))
        output_sheet.delete_cols(column_index_from_string("I"))
        output_sheet.delete_cols(column_index_from_string("E"))

        # 获取 F 列的索引
        column_index = column_index_from_string("F")
        # 反向遍历每一行，如果 F 列无数据或者是 None，则删除该行
        for row in reversed(list(output_sheet.iter_rows(min_row=1, max_row=output_sheet.max_row))):
            if row[column_index - 1].value is None or not row[column_index - 1].value or '单位' in row[
                column_index - 1].value or row[column_index - 1].value == '/':
                output_sheet.delete_rows(row[0].row)
        # 保存新的工作簿
        output_wb.save(f"{process_path}/{index + 1}-{ws['A2'].value.replace('工程名称：', '')}-{category}.xlsx")


def split_excel_danwei(input_file, process_path, category, part):
    wb = pl.load_workbook(input_file, data_only=True)
    for index, sheet_name in enumerate(wb.sheetnames):
        # 创建一个新的工作簿
        output_wb = pl.Workbook()
        # 获取当前工作表
        source_sheet = wb[sheet_name]
        # 在新工作簿中创建相应的工作表
        output_sheet = output_wb.active
        ws = wb[sheet_name]
        output_sheet.title = part  # 以后可以改成sheet_name
        # 复制数据到新的工作表
        for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
            # 检查特定关键词
            keywords_to_skip_danwei = ["规费", "安全生产、文明施工费", "税金"]
            # 如果B列的值不等于关键词，跳过
            if row[1] not in keywords_to_skip_danwei:
                continue
            output_sheet.append(row)
        # 在表的第一列前拆入一列
        output_sheet.insert_cols(1)
        # 将这一列的中单元格数据都填充为ws['A2'].value
        for i in range(2, output_sheet.max_row + 1):
            # 其中的输入都包含"工程名称："，去掉这个字符串
            output_sheet[f'A{i}'] = ws['A2'].value.replace('工程名称：', '')
        # 如果单元格中的值是/，则替换为0
        for row in output_sheet.iter_rows():
            for cell in row:
                if cell.value == "/":
                    cell.value = 0

        output_sheet.delete_cols(column_index_from_string("H"))
        output_sheet.delete_cols(column_index_from_string("H"))
        output_sheet.delete_cols(column_index_from_string("H"))
        output_sheet.delete_cols(column_index_from_string("H"))
        # 保存新的工作簿
        output_wb.save(f"{process_path}/{index + 1}-{ws['A2'].value.replace('工程名称：', '')}-{category}.xlsx")


def split_excel_djcs(input_file, process_path, category, part):
    wb = pl.load_workbook(input_file, data_only=True)
    for index, sheet_name in enumerate(wb.sheetnames):
        # 创建一个新的工作簿
        output_wb = pl.Workbook()
        # 获取当前工作表
        source_sheet = wb[sheet_name]
        # 在新工作簿中创建相应的工作表
        output_sheet = output_wb.active
        ws = wb[sheet_name]
        output_sheet.title = part  # 以后可以改成sheet_name
        # 复制数据到新的工作表
        for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
            output_sheet.append(row)
        # 在表的第一列前拆入一列
        output_sheet.insert_cols(1)
        # 将这一列的中单元格数据都填充为ws['A2'].value
        for i in range(2, output_sheet.max_row + 1):
            # 其中的输入都包含"工程名称："，去掉这个字符串
            output_sheet[f'A{i}'] = ws['A2'].value.replace('工程名称：', '')
        # 删除E列和I列
        # output_sheet.delete_cols(column_index_from_string("H"))
        output_sheet.delete_cols(column_index_from_string("I"))
        output_sheet.delete_cols(column_index_from_string("E"))

        # 获取 F 列的索引
        column_index = column_index_from_string("F")
        # 反向遍历每一行，如果 F 列无数据或者是 None，则删除该行
        for row in reversed(list(output_sheet.iter_rows(min_row=1, max_row=output_sheet.max_row))):
            if row[column_index - 1].value is None or not row[column_index - 1].value or '单位' in row[
                column_index - 1].value or row[column_index - 1].value == '/':
                output_sheet.delete_rows(row[0].row)
        # 保存新的工作簿
        output_wb.save(f"{process_path}/{index + 1}-{ws['A2'].value.replace('工程名称：', '')}-{category}.xlsx")


def split_excel_zjcs(input_file, process_path, category, part):
    wb = pl.load_workbook(input_file, data_only=True)
    for index, sheet_name in enumerate(wb.sheetnames):
        # 创建一个新的工作簿
        output_wb = pl.Workbook()
        # 获取当前工作表
        source_sheet = wb[sheet_name]
        # 在新工作簿中创建相应的工作表
        output_sheet = output_wb.active
        ws = wb[sheet_name]
        output_sheet.title = part  # 以后可以改成sheet_name
        # 复制数据到新的工作表
        for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
            output_sheet.append(row)
        # 在表的第一列前拆入一列
        output_sheet.insert_cols(1)
        # 将这一列的中单元格数据都填充为ws['A2'].value
        for i in range(2, output_sheet.max_row + 1):
            # 其中的输入都包含"工程名称："，去掉这个字符串
            output_sheet[f'A{i}'] = ws['A2'].value.replace('工程名称：', '')
        # 删除D列和F列
        output_sheet.delete_cols(column_index_from_string("F"))
        output_sheet.delete_cols(column_index_from_string("D"))

        for row in reversed(list(output_sheet.iter_rows())):
            cell_value = row[column_index_from_string("D") - 1].value
            if cell_value is None:
                output_sheet.delete_rows(row[0].row)
            elif cell_value == '小计':
                output_sheet.delete_rows(row[0].row)
            elif '安全' in cell_value:
                output_sheet.delete_rows(row[0].row)
            elif cell_value == '项目名称':
                output_sheet.delete_rows(row[0].row)
        # 保存新的工作簿
        output_wb.save(f"{process_path}/{index + 1}-{ws['A2'].value.replace('工程名称：', '')}-{category}.xlsx")


def handle_excel_dx(input_file_name, save_path):
    wb = pl.load_workbook(input_file_name)
    ws = wb.active
    # 倒序读取表格行，如果A列有小数点，则删除该行
    # 反向遍历每一行，如果 F 列无数据或者是 None，则删除该行
    for row in reversed(list(ws.iter_rows(min_row=1, max_row=ws.max_row))):
        if row[0].value is None:
            continue
        if "." in row[0].value or '/' in row[0].value:
            ws.delete_rows(row[0].row)
    ws.delete_cols(column_index_from_string("E"))
    ws.delete_cols(column_index_from_string("E"))
    ws.delete_cols(column_index_from_string("E"))
    ws.delete_cols(column_index_from_string("C"))
    ws.delete_rows(1)
    ws.delete_rows(1)
    ws.delete_rows(1)
    ws.delete_rows(1)
    # B列单元格中的“ 合计”字段删除
    for row in ws.iter_rows():
        if row[1].value is None:
            continue
        row[1].value = row[1].value.replace(" 合计", "")
    filename = os.path.basename(input_file_name)
    saved_file_name = os.path.join(save_path, filename)
    wb.save(saved_file_name)


def fujian4(Contract_Name, Reporting_Periods, year, month, Contract_Number, fenbu_path_ht, djcs_path_ht, zjcs_path_ht, danwei_path_ht,
            fenbu_path_sent, djcs_path_sent, zjcs_path_sent, danwei_path_sent,
            fenbu_path_audited, djcs_path_audited, zjcs_path_audited, danwei_path_audited, output_path,
            input_ht_dx):

    # 获取合同
    # 分部分项工程清单内容
    file_list_fenbu_ht = os.listdir(fenbu_path_ht)
    file_list_fenbu_ht.sort(key=lambda x: int(x.split('-')[0]))
    print(f"file_list_fenbu_ht:{file_list_fenbu_ht}")
    # 获取单位工程费汇总表内容
    file_list_danwei_ht = os.listdir(danwei_path_ht)
    file_list_danwei_ht.sort(key=lambda x: int(x.split('-')[0]))
    print(f"file_list_danwei_ht:{file_list_danwei_ht}")
    # 获取单价措施内容
    file_list_djcs_ht = os.listdir(djcs_path_ht)
    file_list_djcs_ht.sort(key=lambda x: int(x.split('-')[0]))
    print(f'file_list_djcs_ht:{file_list_djcs_ht}')
    # 获取总价措施内容
    file_list_zjcs_ht = os.listdir(zjcs_path_ht)
    file_list_zjcs_ht.sort(key=lambda x: int(x.split('-')[0]))
    print(f'file_list_zjcs_ht:{file_list_zjcs_ht}')

    # 获取送审
    # 分部分项工程清单内容
    file_list_fenbu_sent = os.listdir(fenbu_path_sent)
    file_list_fenbu_sent.sort(key=lambda x: int(x.split('-')[0]))
    print(f"file_list_fenbu_sent:{file_list_fenbu_sent}")
    # 获取单位工程费汇总表内容
    file_list_danwei_sent = os.listdir(danwei_path_sent)
    file_list_danwei_sent.sort(key=lambda x: int(x.split('-')[0]))
    print(f"file_list_danwei_sent:{file_list_danwei_sent}")
    # 获取单价措施内容
    file_list_djcs_sent = os.listdir(djcs_path_sent)
    file_list_djcs_sent.sort(key=lambda x: int(x.split('-')[0]))
    print(f'file_list_djcs_sent:{file_list_djcs_sent}')
    # 获取总价措施内容
    file_list_zjcs_sent = os.listdir(zjcs_path_sent)
    file_list_zjcs_sent.sort(key=lambda x: int(x.split('-')[0]))
    print(f'file_list_zjcs_sent:{file_list_zjcs_sent}')

    # 获取审定
    # 分部分项工程清单内容
    file_list_fenbu_audited = os.listdir(fenbu_path_audited)
    file_list_fenbu_audited.sort(key=lambda x: int(x.split('-')[0]))
    print(f"file_list_fenbu_audited:{file_list_fenbu_audited}")
    # 获取单位工程费汇总表内容
    file_list_danwei_audited = os.listdir(danwei_path_audited)
    file_list_danwei_audited.sort(key=lambda x: int(x.split('-')[0]))
    print(f"file_list_danwei_audited:{file_list_danwei_audited}")
    # 获取单价措施内容
    file_list_djcs_audited = os.listdir(djcs_path_audited)
    file_list_djcs_audited.sort(key=lambda x: int(x.split('-')[0]))
    print(f'file_list_djcs_audited:{file_list_djcs_audited}')
    # 获取总价措施内容
    file_list_zjcs_audited = os.listdir(zjcs_path_audited)
    file_list_zjcs_audited.sort(key=lambda x: int(x.split('-')[0]))
    print(f'file_list_zjcs_audited:{file_list_zjcs_audited}')
    #
    # 获取模板文件
    file_muban_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "function", "progress_reports_xd")
    file_name = "模板.xlsx"
    file_mb = os.path.join(file_muban_path, file_name)
    wb_muban = pl.load_workbook(file_mb, data_only=True)
    ws_muban_fujian4 = wb_muban.active
    # 清除后面所有的数据
    for row in ws_muban_fujian4.iter_rows():
        # 跳过第一行至第五行，从第六行开始
        if row[0].row < 6:
            continue
        # 清除后面所有的数据
        for cell in row:
            cell.value = None
    project_name_list = []
    project_name_row_list = []
    for (fenbu_ht, djcs_ht, zjcs_ht, danwei_ht,
         fenbu_sent, djcs_sent, zjcs_sent, danwei_sent,
         fenbu_audited, djcs_audited, zjcs_audited, danwei_audited) in zip(
        file_list_fenbu_ht, file_list_djcs_ht, file_list_zjcs_ht, file_list_danwei_ht,
        file_list_fenbu_sent, file_list_djcs_sent, file_list_zjcs_sent, file_list_danwei_sent,
        file_list_fenbu_audited, file_list_djcs_audited, file_list_zjcs_audited, file_list_danwei_audited):

        # 合 同
        # 分部分项
        wb_fenbu_ht = pl.load_workbook(os.path.join(os.path.dirname(__file__), fenbu_path_ht, fenbu_ht), data_only=True)
        ws_fenbu_ht = wb_fenbu_ht.active
        # 单价措施
        wb_djcs_ht = pl.load_workbook(os.path.join(os.path.dirname(__file__), djcs_path_ht, djcs_ht), data_only=True)
        ws_djcs_ht = wb_djcs_ht.active
        # 总价措施
        wb_zjcs_ht = pl.load_workbook(os.path.join(os.path.dirname(__file__), zjcs_path_ht, zjcs_ht), data_only=True)
        ws_zjcs_ht = wb_zjcs_ht.active
        # 单位工程
        wb_danwei_ht = pl.load_workbook(os.path.join(os.path.dirname(__file__), danwei_path_ht, danwei_ht),
                                        data_only=True)
        ws_danwei_ht = wb_danwei_ht.active
        # 送 审
        # 分部分项
        wb_fenbu_sent = pl.load_workbook(os.path.join(os.path.dirname(__file__), fenbu_path_sent, fenbu_sent),
                                         data_only=True)
        ws_fenbu_sent = wb_fenbu_sent.active
        # 单价措施
        wb_djcs_sent = pl.load_workbook(os.path.join(os.path.dirname(__file__), djcs_path_sent, djcs_sent),
                                        data_only=True)
        ws_djcs_sent = wb_djcs_sent.active
        # 总价措施
        wb_zjcs_sent = pl.load_workbook(os.path.join(os.path.dirname(__file__), zjcs_path_sent, zjcs_sent),
                                        data_only=True)
        ws_zjcs_sent = wb_zjcs_sent.active
        # 单位工程
        wb_danwei_sent = pl.load_workbook(os.path.join(os.path.dirname(__file__), danwei_path_sent, danwei_sent),
                                          data_only=True)
        ws_danwei_sent = wb_danwei_sent.active
        # 审 定
        # 分部分项
        wb_fenbu_audited = pl.load_workbook(os.path.join(os.path.dirname(__file__), fenbu_path_audited, fenbu_audited),
                                            data_only=True)
        ws_fenbu_audited = wb_fenbu_audited.active
        # 单价措施
        wb_djcs_audited = pl.load_workbook(os.path.join(os.path.dirname(__file__), djcs_path_audited, djcs_audited),
                                           data_only=True)
        ws_djcs_audited = wb_djcs_audited.active
        # 总价措施
        wb_zjcs_audited = pl.load_workbook(os.path.join(os.path.dirname(__file__), zjcs_path_audited, zjcs_audited),
                                           data_only=True)
        ws_zjcs_audited = wb_zjcs_audited.active
        # 单位工程
        wb_danwei_audited = pl.load_workbook(
            os.path.join(os.path.dirname(__file__), danwei_path_audited, danwei_audited), data_only=True)
        ws_danwei_audited = wb_danwei_audited.active

        # 获取 fenbu_ht 的文件名称
        project_name = os.path.basename(fenbu_ht).split('-', 1)[1].strip('.xlsx')
        print(f"工程名称：{project_name}")


        # 把工程名称写入附件4的C列
        write_col_project_name = ws_muban_fujian4.max_row + 1
        write_col = ws_muban_fujian4.max_row + 1
        ws_muban_fujian4[f'A{write_col}'] = write_col - 5
        ws_muban_fujian4[f'C{write_col}'] = project_name
        project_name_list.append(project_name)
        project_name_row_list.append(write_col)
        write_col = ws_muban_fujian4.max_row + 1
        fbfx_name = [write_col - 5, None, "分部分项工程量清单部分", None, None, None, None, None, None, None, None,
                     None, None]
        ws_muban_fujian4.append(fbfx_name)
        # print(f"分部分项工程量清单部分{fenbu_ht, fenbu_sent, fenbu_audited}")
        for row_ht, row_sent, row_audited in zip(
                ws_fenbu_ht.iter_rows(max_col=ws_fenbu_ht.max_column + 10, values_only=True),
                ws_fenbu_sent.iter_rows(max_col=ws_fenbu_sent.max_column + 10, values_only=True),
                ws_fenbu_audited.iter_rows(max_col=ws_fenbu_audited.max_column + 10, values_only=True)):
            print(f"分部{fenbu_ht, fenbu_sent, fenbu_audited}")

            bianhao = row_ht[column_index_from_string("B") - 1]
            name = row_ht[column_index_from_string("C") - 1]
            danwei = row_ht[column_index_from_string("E") - 1]
            gongchengliang = row_ht[column_index_from_string("F") - 1]
            danjia = row_ht[column_index_from_string("G") - 1]
            if not danjia or danjia is None:
                danjia = 0
            danjia = float(danjia)
            hejia = row_ht[column_index_from_string("H") - 1]
            if not hejia or hejia is None:
                hejia = 0
            gongchengliang_sent = row_sent[column_index_from_string("F") - 1]
            if gongchengliang_sent is None:
                gongchengliang_sent = 0
            gongchengliang_sent = float(gongchengliang_sent)
            hejia_sent = gongchengliang_sent * danjia
            if hejia_sent is None:
                hejia_sent = 0
            gongchengliang_audited = row_audited[column_index_from_string("F") - 1]
            if gongchengliang_audited is None:
                gongchengliang_audited = 0
            gongchengliang_audited = float(gongchengliang_audited)
            hejia_audited = gongchengliang_audited * danjia
            print(
                f"分部分项值:{bianhao, name, danwei, gongchengliang, danjia, hejia, gongchengliang_sent, hejia_sent, gongchengliang_audited, hejia_audited}")

            write_col = ws_muban_fujian4.max_row + 1
            ws_muban_fujian4[f'A{write_col}'] = write_col - 5
            ws_muban_fujian4[f'B{write_col}'] = bianhao
            ws_muban_fujian4[f'C{write_col}'] = name
            ws_muban_fujian4[f'D{write_col}'] = danwei
            ws_muban_fujian4[f'E{write_col}'] = gongchengliang
            ws_muban_fujian4[f'F{write_col}'] = danjia
            ws_muban_fujian4[f'G{write_col}'] = hejia
            ws_muban_fujian4[f'H{write_col}'] = None
            ws_muban_fujian4[f'I{write_col}'] = None
            ws_muban_fujian4[f'J{write_col}'] = gongchengliang_sent
            ws_muban_fujian4[f'K{write_col}'] = hejia_sent
            ws_muban_fujian4[f'L{write_col}'] = gongchengliang_audited
            ws_muban_fujian4[f'M{write_col}'] = hejia_audited

        write_col = ws_muban_fujian4.max_row + 1
        djcs_name = [write_col - 5, None, "单价措施", None, None, None, None, None, None, None, None, None,
                     None]
        ws_muban_fujian4.append(djcs_name)
        for row_ht, row_sent, row_audited in zip(
                ws_djcs_ht.iter_rows(values_only=True, max_col=ws_djcs_ht.max_column + 10),
                ws_djcs_sent.iter_rows(values_only=True, max_col=ws_djcs_sent.max_column + 10),
                ws_djcs_audited.iter_rows(values_only=True, max_col=ws_djcs_audited.max_column + 10)):

            print(f'单价措施{djcs_ht, djcs_sent, djcs_audited, djcs_path_sent}')
            danjia = row_ht[column_index_from_string("G") - 1]
            if danjia is None:
                pass
            else:
                bianhao = row_ht[column_index_from_string("B") - 1]
                name = row_ht[column_index_from_string("C") - 1]
                danwei = row_ht[column_index_from_string("E") - 1]
                gongchengliang = row_ht[column_index_from_string("F") - 1]
                danjia = row_ht[column_index_from_string("G") - 1]
                danjia = float(danjia)
                hejia = row_ht[column_index_from_string("H") - 1]
                gongchengliang_sent = row_sent[column_index_from_string("F") - 1]
                if gongchengliang_sent is None:
                    gongchengliang_sent = 0
                gongchengliang_sent = float(gongchengliang_sent)
                hejia_sent = gongchengliang_sent * danjia
                gongchengliang_audited = row_audited[column_index_from_string("F") - 1]
                if gongchengliang_audited is None:
                    gongchengliang_audited = 0
                gongchengliang_audited = float(gongchengliang_audited)
                hejia_audited = gongchengliang_audited * danjia
                print(
                    f"单价措施-值{bianhao, name, danwei, gongchengliang, danjia, hejia, gongchengliang_sent, hejia_sent, gongchengliang_audited, hejia_audited}")
                write_col = ws_muban_fujian4.max_row + 1
                ws_muban_fujian4[f'A{write_col}'] = write_col - 5
                ws_muban_fujian4[f'B{write_col}'] = bianhao
                ws_muban_fujian4[f'C{write_col}'] = name
                ws_muban_fujian4[f'D{write_col}'] = danwei
                ws_muban_fujian4[f'E{write_col}'] = gongchengliang
                ws_muban_fujian4[f'F{write_col}'] = danjia
                ws_muban_fujian4[f'G{write_col}'] = hejia
                ws_muban_fujian4[f'H{write_col}'] = None
                ws_muban_fujian4[f'I{write_col}'] = None
                ws_muban_fujian4[f'J{write_col}'] = gongchengliang_sent
                ws_muban_fujian4[f'K{write_col}'] = hejia_sent
                ws_muban_fujian4[f'L{write_col}'] = gongchengliang_audited
                ws_muban_fujian4[f'M{write_col}'] = hejia_audited
        write_col = ws_muban_fujian4.max_row + 1
        zongjiacuoshi_name = [write_col - 5, None, "其他总价措施", None, None, None, None, None, None, None, None, None,
                              None]
        ws_muban_fujian4.append(zongjiacuoshi_name)
        for row_ht, row_sent, row_audited in zip(
                ws_zjcs_ht.iter_rows(values_only=True, max_col=ws_zjcs_ht.max_column + 10),
                ws_zjcs_sent.iter_rows(values_only=True, max_col=ws_zjcs_sent.max_column + 10),
                ws_zjcs_audited.iter_rows(values_only=True, max_col=ws_zjcs_audited.max_column + 10)):
            # print(f"总价措施{zjcs_ht, zjcs_sent, zjcs_audited}")
            bianhao = row_ht[column_index_from_string("B") - 1]
            name = row_ht[column_index_from_string("C") - 1]
            hejia = row_ht[column_index_from_string("D") - 1]
            hejia_sent = row_sent[column_index_from_string("D") - 1]
            hejia_audited = row_audited[column_index_from_string("D") - 1]
            # print(f"values:{bianhao, name, hejia, hejia_sent, hejia_audited}")
            write_col = ws_muban_fujian4.max_row + 1
            ws_muban_fujian4[f'A{write_col}'] = write_col - 5
            ws_muban_fujian4[f'B{write_col}'] = bianhao
            ws_muban_fujian4[f'C{write_col}'] = name
            ws_muban_fujian4[f'D{write_col}'] = None
            ws_muban_fujian4[f'E{write_col}'] = None
            ws_muban_fujian4[f'F{write_col}'] = None
            ws_muban_fujian4[f'G{write_col}'] = hejia
            ws_muban_fujian4[f'H{write_col}'] = None
            ws_muban_fujian4[f'I{write_col}'] = None
            ws_muban_fujian4[f'J{write_col}'] = None
            ws_muban_fujian4[f'K{write_col}'] = hejia_sent
            ws_muban_fujian4[f'L{write_col}'] = None
            ws_muban_fujian4[f'M{write_col}'] = hejia_audited
        for row_ht, row_sent, row_audited in zip(
                ws_danwei_ht.iter_rows(values_only=True, max_col=ws_danwei_ht.max_column + 10),
                ws_danwei_sent.iter_rows(values_only=True, max_col=ws_danwei_sent.max_column + 10),
                ws_danwei_audited.iter_rows(values_only=True, max_col=ws_danwei_audited.max_column + 10)):
            # print(f"单位工程{danwei_ht, danwei_sent, danwei_audited}")
            name = row_ht[column_index_from_string("B") - 1]
            price = row_ht[column_index_from_string("E") - 1]
            price_sent = row_sent[column_index_from_string("E") - 1]
            price_audited = row_audited[column_index_from_string("E") - 1]
            write_col = ws_muban_fujian4.max_row + 1
            ws_muban_fujian4[f'A{write_col}'] = write_col - 5
            ws_muban_fujian4[f'B{write_col}'] = None
            ws_muban_fujian4[f'C{write_col}'] = name
            ws_muban_fujian4[f'D{write_col}'] = None
            ws_muban_fujian4[f'E{write_col}'] = None
            ws_muban_fujian4[f'F{write_col}'] = None
            ws_muban_fujian4[f'G{write_col}'] = price
            ws_muban_fujian4[f'H{write_col}'] = None
            ws_muban_fujian4[f'I{write_col}'] = None
            ws_muban_fujian4[f'J{write_col}'] = None
            ws_muban_fujian4[f'K{write_col}'] = price_sent
            ws_muban_fujian4[f'L{write_col}'] = None
            ws_muban_fujian4[f'M{write_col}'] = price_audited
        # break
        write_col = ws_muban_fujian4.max_row

        sum_value_ht = 0
        sum_value_send = 0
        sum_value_audited = 0
        sum_value_shangqi = 0
        sum_value_benqi_duibi = 0
        sum_value_benqi_leiji = 0
        sum_value_leiji = 0
        for row in ws_muban_fujian4.iter_rows(min_row=write_col_project_name + 1, max_row=write_col):
            # 跳过前 5 行
            if row[0].row < 6:
                continue
            value_ht = row[column_index_from_string("G") - 1].value
            value_send = row[column_index_from_string("K") - 1].value
            value_audited = row[column_index_from_string("M") - 1].value
            value_shangqi = row[column_index_from_string("I") - 1].value
            value_benqi_duibi = row[column_index_from_string("O") - 1].value
            value_benqi_leiji = row[column_index_from_string("Q") - 1].value
            value_leiji = row[column_index_from_string("R") - 1].value
            if value_ht is None or value_ht == '':
                value_ht = 0
            if value_send is None or value_send == '':
                value_send = 0
            if value_audited is None or value_audited == '':
                value_audited = 0
            if value_shangqi is None or value_shangqi == '':
                value_shangqi = 0
            if value_benqi_duibi is None or value_benqi_duibi == '':
                value_benqi_duibi = 0
            if value_benqi_leiji is None or value_benqi_leiji == '':
                value_benqi_leiji = 0
            if value_leiji is None or value_leiji == '':
                value_leiji = 0
            # print(row[0].row)
            value_ht = float(value_ht)
            value_send = float(value_send)
            value_audited = float(value_audited)
            value_shangqi = float(value_shangqi)
            value_benqi_duibi = float(value_benqi_duibi)
            value_benqi_leiji = float(value_benqi_leiji)
            value_leiji = float(value_leiji)

            sum_value_ht += value_ht
            sum_value_send += value_send
            sum_value_audited += value_audited
            sum_value_shangqi += value_shangqi
            sum_value_benqi_duibi += value_benqi_duibi
            sum_value_benqi_leiji += value_benqi_leiji
            sum_value_leiji += value_leiji
        ws_muban_fujian4[f'G{write_col_project_name}'].value = sum_value_ht
        ws_muban_fujian4[f'K{write_col_project_name}'].value = sum_value_send
        ws_muban_fujian4[f'M{write_col_project_name}'].value = sum_value_audited
        ws_muban_fujian4[f'I{write_col_project_name}'].value = sum_value_shangqi
        ws_muban_fujian4[f'O{write_col_project_name}'].value = sum_value_benqi_duibi
        ws_muban_fujian4[f'Q{write_col_project_name}'].value = sum_value_benqi_leiji
        ws_muban_fujian4[f'R{write_col_project_name}'].value = sum_value_leiji

    e_num = column_index_from_string("E")
    r_num = column_index_from_string("R")
    for row in ws_muban_fujian4.iter_rows(min_row=6, min_col=e_num, max_col=r_num):
        for cell in row:
            # 将字符串转换为整数并更新单元格的值
            if cell.value is None or cell.value == '':
                continue
            cell.value = float(cell.value)

    # 计算本期对比及本期末累计完成，及累计金额对比
    # 本期对比，本期对比中有工程量和金额的对比，需要分别计算
    for row in ws_muban_fujian4.iter_rows(min_row=6, max_row=ws_muban_fujian4.max_row):
        gcl_sent_num = column_index_from_string("J") - 1
        gcl_audited_num = column_index_from_string("L") - 1
        price_sent_num = column_index_from_string("K") - 1
        price_audited_num = column_index_from_string("M") - 1

        gcl_sent = row[gcl_sent_num].value
        gcl_audited = row[gcl_audited_num].value
        price_sent = row[price_sent_num].value
        price_audited = row[price_audited_num].value

        write_col = row[0].row
        if gcl_sent is None or gcl_sent == '':
            gcl_sent = 0
        if gcl_audited is None or gcl_audited == '':
            gcl_audited = 0
        if price_sent is None or price_sent == '':
            price_sent = 0
        if price_audited is None or price_audited == '':
            price_audited = 0
        if gcl_audited - gcl_sent == 0:
            ws_muban_fujian4[f'N{write_col}'] = ""
        else:
            ws_muban_fujian4[f'N{write_col}'] = gcl_audited - gcl_sent
        if price_audited - price_sent == 0:
            ws_muban_fujian4[f'O{write_col}'] = ""
        else:
            ws_muban_fujian4[f'O{write_col}'] = price_audited - price_sent

    # 本期末累计完成
    for row in ws_muban_fujian4.iter_rows(min_row=6, max_row=ws_muban_fujian4.max_row):
        gcl_num = column_index_from_string("H") - 1
        price_num = column_index_from_string("I") - 1
        gcl_audited_num = column_index_from_string("L") - 1
        price_audited_num = column_index_from_string("M") - 1
        gcl = row[gcl_num].value
        price = row[price_num].value
        gcl_audited = row[gcl_audited_num].value
        price_audited = row[price_audited_num].value
        write_col = row[0].row
        if gcl is None or gcl == '':
            gcl = 0
        if price is None or price == '':
            price = 0
        if gcl_audited is None or gcl_audited == '':
            gcl_audited = 0
        if price_audited is None or price_audited == '':
            price_audited = 0
        if gcl + gcl_audited == 0:
            ws_muban_fujian4[f'P{write_col}'] = ""
        else:
            ws_muban_fujian4[f'P{write_col}'] = gcl + gcl_audited
        if price + price_audited == 0:
            ws_muban_fujian4[f'Q{write_col}'] = ""
        else:
            ws_muban_fujian4[f'Q{write_col}'] = price + price_audited

    shangqi_price_dict = {}
    sent_price_dict = {}
    audited_price_dict = {}
    duibi_price_dict = {}
    leiji_price_dict = {}
    for row in ws_muban_fujian4.iter_rows(min_row=6, max_row=ws_muban_fujian4.max_row):
        name = row[column_index_from_string("C") - 1].value
        shangqi_price = row[column_index_from_string("I") - 1].value
        sent_price = row[column_index_from_string("K") - 1].value
        audited_price = row[column_index_from_string("M") - 1].value
        duibi_price = row[column_index_from_string("O") - 1].value
        leiji_price = row[column_index_from_string("Q") - 1].value
        shangqi_price_dict[name] = shangqi_price
        sent_price_dict[name] = sent_price
        audited_price_dict[name] = audited_price
        duibi_price_dict[name] = duibi_price
        leiji_price_dict[name] = leiji_price

    shangqi_price = 0
    sent_price = 0
    audited_price = 0
    duibi_price = 0
    leiji_price = 0
    for name in project_name_list:
        if shangqi_price_dict[name] is None or shangqi_price_dict[name] == '':
            shangqi_price_dict[name] = 0
        if sent_price_dict[name] is None or sent_price_dict[name] == '':
            sent_price_dict[name] = 0
        if audited_price_dict[name] is None or audited_price_dict[name] == '':
            audited_price_dict[name] = 0
        if duibi_price_dict[name] is None or duibi_price_dict[name] == '':
            duibi_price_dict[name] = 0
        if leiji_price_dict[name] is None or leiji_price_dict[name] == '':
            leiji_price_dict[name] = 0
        shangqi_price += float(shangqi_price_dict[name])
        sent_price += float(sent_price_dict[name])
        audited_price += float(audited_price_dict[name])
        duibi_price += float(duibi_price_dict[name])
        leiji_price += float(leiji_price_dict[name])

    # 获取合同金额
    for file in os.listdir(input_ht_dx):
        if file.endswith(".xls"):
            excel_name_ht_dx = os.path.join(input_ht_dx, file)
    convert_xls_to_xlsx(xls_filename=excel_name_ht_dx, xlsx_filename=os.path.join(input_ht_dx, "ht_dx.xlsx"),
                        save_path=input_ht_dx)
    wb_dx = pl.load_workbook(os.path.join(input_ht_dx, "ht_dx.xlsx"), data_only=True)
    ws_dx = wb_dx.active
    ws_dx_max_row = ws_dx.max_row
    price_sum_ht = ws_dx[f"D{ws_dx_max_row}"].value
    print(f"合同总价：{price_sum_ht}，此数值来源于D{ws_dx_max_row}单元格")
    write_col_sum = ws_muban_fujian4.max_row + 1
    ws_muban_fujian4[f'A{write_col_sum}'] = write_col_sum - 5
    ws_muban_fujian4[f'C{write_col_sum}'] = '合    计'
    ws_muban_fujian4[f'G{write_col_sum}'] = price_sum_ht
    ws_muban_fujian4[f'I{write_col_sum}'] = shangqi_price
    ws_muban_fujian4[f'K{write_col_sum}'] = sent_price
    ws_muban_fujian4[f'M{write_col_sum}'] = audited_price
    ws_muban_fujian4[f'O{write_col_sum}'] = duibi_price
    ws_muban_fujian4[f'Q{write_col_sum}'] = leiji_price

    # 累计金额对比  本期末累计金额 - 合同总价
    for row in ws_muban_fujian4.iter_rows(min_row=6, max_row=ws_muban_fujian4.max_row):
        price_num = column_index_from_string("Q") - 1
        price_sum_num = column_index_from_string("G") - 1
        price = row[price_num].value
        price_sum = row[price_sum_num].value
        write_col = row[0].row
        if price is None or price == '':
            price = 0
        if price_sum is None or price_sum == '':
            price_sum = 0
        price = float(price)
        price_sum = float(price_sum)
        if price - price_sum == 0:
            ws_muban_fujian4[f'R{write_col}'] = ""
        else:
            ws_muban_fujian4[f'R{write_col}'] = price - price_sum

    e_num = column_index_from_string("E")
    r_num = column_index_from_string("R")
    for row in ws_muban_fujian4.iter_rows(min_row=6, min_col=e_num, max_col=r_num):
        for cell in row:
            # 将字符串转换为整数并更新单元格的值
            if cell.value is None or cell.value == '':
                continue
            cell.value = float(cell.value)

    for row in ws_muban_fujian4.iter_rows():
        if row[0].row < 6:
            continue
        for cell in row:
            cell.font = Font(name='宋体', size=9)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in ws_muban_fujian4.iter_rows(max_col=column_index_from_string("C"), min_col=column_index_from_string("C")):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 设定列宽
    ws_muban_fujian4.column_dimensions['A'].width = 3.5
    ws_muban_fujian4.column_dimensions['B'].width = 13
    ws_muban_fujian4.column_dimensions['C'].width = 30

    for num in project_name_row_list:
        for row in ws_muban_fujian4.iter_rows(min_row=num, max_row=num):
            for cell in row:
                cell.font = Font(bold=True, size=10, name='宋体')
    ws_muban_fujian4[f"C{ws_muban_fujian4.max_row}"].font = Font(bold=True, size=10, name='宋体')

    # 设置表头
    ws_muban_fujian4['A2'] = f'单位工程审核明细对比表（第 {Reporting_Periods} 次进度付款）'
    ws_muban_fujian4['A3'] = f'合同名称：{Contract_Name}' + " " * 30 + f"合同编号：{Contract_Number}" + " " * 30 + f"月份：{year}年第{month}月"
    ws_muban_fujian4['J4'] = f'本期申报（第{Reporting_Periods}次）'
    ws_muban_fujian4['L4'] = f'本期审核（第{Reporting_Periods}次）'
    print(ws_muban_fujian4['J4'].value, ws_muban_fujian4['L4'].value)
    fj_4_name = f'{Contract_Name}第{Reporting_Periods}期进度款单位工程审核明细对比表（附件四）.xlsx'
    save_path = os.path.join(output_path, fj_4_name)
    wb_muban.save(save_path)
    return save_path


def crate_dir(path):
    try:
        os.mkdir(path, 0o777)
    except FileExistsError:
        pass


def get_file_path(input_folder, process_folder_path, leixing):
    # 创建文件目录
    crate_dir(os.path.join(process_folder_path, "ht"))
    crate_dir(os.path.join(process_folder_path, "ss"))
    crate_dir(os.path.join(process_folder_path, "sd"))
    jdk_path = process_folder_path
    for folder in os.listdir(jdk_path):
        # print(folder)
        path_fb = os.path.join(jdk_path, folder, "fb")
        path_dj = os.path.join(jdk_path, folder, "dj")
        path_zj = os.path.join(jdk_path, folder, "zj")
        path_dw = os.path.join(jdk_path, folder, "dw")
        crate_dir(path_fb), crate_dir(path_dj), crate_dir(path_zj), crate_dir(path_dw)
    folder_dir = input_folder
    # 获取folder_dir目录下的所有文件夹, 不包括文件
    folder_1_list = [folder for folder in os.listdir(folder_dir) if os.path.isdir(os.path.join(folder_dir, folder))]
    # 让文件夹按照数字顺序排列
    folder_1_list.sort(key=lambda x: int(x.split(" ")[0]))
    # print(folder_1_list)
    folder_2_path_list = []
    for folder in folder_1_list:
        folder2_dir = os.path.join(folder_dir, folder)
        folder_list = [folder for folder in os.listdir(folder2_dir) if os.path.isdir(os.path.join(folder2_dir, folder))]
        folder_list.sort(key=lambda x: int(x.split(" ")[0]))
        for folder1 in folder_list:
            path = os.path.join(folder2_dir, folder1)
            folder_2_path_list.append(path)

    file_list = []  # 用于存放所有文件的路径
    for file in folder_2_path_list:
        # print(file)
        for f in os.listdir(file):
            file_list.append(os.path.join(file, f))

    for index, file in enumerate(file_list):
        num = index + 1
        base_path = input_folder
        relative_path = os.path.relpath(file, base_path)
        folder_name = os.path.dirname(os.path.dirname(relative_path))
        folder_name.split()
        name_part = [part for part in folder_name.split() if not part.isdigit()]
        process_name = " ".join(name_part)
        file_name_xls = os.path.basename(file)
        # 将文件名称的后缀由xls改为xlsx
        file_name_xlsx = file_name_xls.replace("xls", "xlsx")
        new_file_name = process_name + "-" + file_name_xlsx
        wb = xlrd.open_workbook(file)
        # 获取sheet表格的个数和名称
        sheet_names = wb.sheet_names()
        for sheet_name in sheet_names:
            sheet = wb.sheet_by_name(sheet_name)
            wb_new = pl.Workbook()
            sheet_new = wb_new.active
            sheet_new.title = "sheet"
            for row_index in range(sheet.nrows):
                for col_index in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_index, col_index)
                    sheet_new.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
            if "分部分项" in sheet_name:
                new_file_path_fb = os.path.join(process_folder_path, leixing, "fb", (str(num) + "-" + new_file_name))
                wb_new.save(new_file_path_fb)
                # print(f"{new_file_path_fb}已完成")
            elif "总价措施" in sheet_name:
                new_file_path_zj = os.path.join(process_folder_path, leixing, "zj", (str(num) + "-" + new_file_name))
                wb_new.save(new_file_path_zj)
                # print(f"{new_file_path_zj}已完成")
            elif "单价措施" in sheet_name:
                new_file_path_dw = os.path.join(process_folder_path, leixing, "dj", (str(num) + "-" + new_file_name))
                wb_new.save(new_file_path_dw)
                # print(f"{new_file_path_dw}已完成")
            elif "单位工程费汇总" in sheet_name:
                new_file_path_dw = os.path.join(process_folder_path, leixing, "dw", (str(num) + "-" + new_file_name))
                wb_new.save(new_file_path_dw)
                # print(f"{new_file_path_dw}已完成")


def fb_excel(excel_path):
    file_list = os.listdir(excel_path)
    file_list.sort(key=lambda x: int(x.split("-")[0]))
    for index, file in enumerate(file_list):
        excel_name = os.path.join(excel_path, file)
        wb = pl.load_workbook(excel_name, data_only=True)
        # print(index + 1, wb.sheetnames)
        ws = wb.active
        ws.delete_cols(column_index_from_string("H"))
        ws.delete_cols(column_index_from_string("D"))
        column_index = column_index_from_string("E")
        # 反向遍历每一行，如果 F 列无数据或者是 None，则删除该行
        for row in reversed(list(ws.iter_rows(min_row=1, max_row=ws.max_row))):
            if row[column_index - 1].value is None or not row[column_index - 1].value or '单位' in row[column_index - 1].value or row[column_index - 1].value == '/':
                ws.delete_rows(row[0].row)
        wb.save(excel_name)


def dj_excel(excel_path):
    file_list = os.listdir(excel_path)
    file_list.sort(key=lambda x: int(x.split("-")[0]))
    for index, file in enumerate(file_list):
        excel_name = os.path.join(excel_path, file)
        wb = pl.load_workbook(excel_name, data_only=True)
        # `print`(index + 1, wb.sheetnames)
        ws = wb.active

        ws.delete_cols(column_index_from_string("H"))
        ws.delete_cols(column_index_from_string("D"))

        # 获取 F 列的索引
        column_index = column_index_from_string("E")
        # 反向遍历每一行，如果 F 列无数据或者是 None，则删除该行
        for row in reversed(list(ws.iter_rows(min_row=1, max_row=ws.max_row))):
            if row[column_index - 1].value is None or not row[column_index - 1].value or '单位' in row[column_index - 1].value or row[column_index - 1].value == '/':
                ws.delete_rows(row[0].row)
        wb.save(excel_name)


def zj_excel(excel_path):
    file_list = os.listdir(excel_path)
    # 忽略掉隐藏文件
    file_list = [file for file in file_list if not file.startswith(".")]
    file_list.sort(key=lambda x: int(x.split("-")[0]))
    for index, file in enumerate(file_list):
        excel_name = os.path.join(excel_path, file)
        wb = pl.load_workbook(excel_name, data_only=True)
        # print(index + 1, wb.sheetnames)
        ws = wb.active

        ws.delete_cols(column_index_from_string("E"))
        ws.delete_cols(column_index_from_string("C"))

        for row in reversed(list(ws.iter_rows())):
            cell_value = row[column_index_from_string("C") - 1].value
            if cell_value is None:
                ws.delete_rows(row[0].row)
            elif cell_value == '小计':
                ws.delete_rows(row[0].row)
            elif '安全' in cell_value:
                ws.delete_rows(row[0].row)
            elif cell_value == '项目名称':
                ws.delete_rows(row[0].row)
        wb.save(excel_name)


def dw_excel(excel_path):
    file_list = os.listdir(excel_path)
    file_list.sort(key=lambda x: int(x.split("-")[0]))
    for index, file in enumerate(file_list):
        excel_name = os.path.join(excel_path, file)
        wb = pl.load_workbook(excel_name, data_only=True)
        # print(index + 1, wb.sheetnames)
        ws = wb.active
        # 依次删除 GHIG 和 C 列
        ws.delete_cols(column_index_from_string("G"))
        ws.delete_cols(column_index_from_string("G"))
        ws.delete_cols(column_index_from_string("G"))
        ws.delete_cols(column_index_from_string("G"))
        ws.delete_cols(column_index_from_string("C"))
        # 获取 F 列的索引
        column_index = column_index_from_string("B")
        keyword_list = ['规费', "安全生产、文明施工费", "税金"]
        # 反向遍历每一行，如果 F 列无数据或者是 None，则删除该行
        for row in reversed(list(ws.iter_rows(min_row=1, max_row=ws.max_row))):
            if row[column_index - 1].value not in keyword_list:
                ws.delete_rows(row[0].row)
        wb.save(excel_name)

# ————————————生成附件3的操作——————————————————
def create_fujian3(input_path: str, output_path: str, Contract_Name, Contract_Number, Reporting_Periods, year, month):
    # 检查是否有历史数据，如果有，则复用历史数据



    for leixing in ["ht", "ss", "sd"]:
        for file in os.listdir(os.path.join(input_path, leixing)):
            # print(file)
            # 判断如果是文件夹就跳过
            if os.path.isdir(os.path.join(input_path, leixing, file)):
                continue
            if file.startswith("."):
                continue
            if file.endswith(".xls"):
                excel_name = os.path.join(input_path, leixing, file)
                save_path = os.path.join(input_path, leixing)
                convert_xls_to_xlsx(xls_filename=excel_name, xlsx_filename=f"{leixing}_dx.xlsx",
                                    save_path=os.path.join(input_path, leixing))
            wb = pl.load_workbook(os.path.join(input_path, leixing, f"{leixing}_dx.xlsx"), data_only=True)
            ws = wb.active
            merged_cells = ws.merged_cells.ranges
            for merged_cell in merged_cells:
                ws.unmerge_cells(str(merged_cell))
            # 倒序遍历每一行，如果第一列中有".", 则删除该行
            for row in ws.iter_rows():
                if row[0].value is None or row[0].value == "":
                    ws.delete_rows(row[0].row)
            for row in reversed(list(ws.iter_rows())):
                if row[0].value is None or row[0].value == "":
                    continue
                if "." in row[0].value or "序号" in row[0].value or "工程" in row[0].value or "/" in row[
                    0].value or "单项工程" in row[0].value or "规费" in row[column_index_from_string("E") - 1].value:
                    ws.delete_rows(row[0].row)
            for row in ws.iter_rows():
                if row[0].value is None or row[0].value == "":
                    ws.delete_rows(row[0].row)

            ws.delete_cols(column_index_from_string("E"))
            ws.delete_cols(column_index_from_string("E"))
            ws.delete_cols(column_index_from_string("E"))
            ws.delete_cols(column_index_from_string("C"))

            for row in ws.iter_rows():
                try:
                    row[1].value = row[1].value.replace(" 合计", "")
                except AttributeError:  # 如果row[1].value是None，则跳过
                    print(row[1].value)
                    pass
            wb.save(os.path.join(input_path, leixing, f"{leixing}_dx.xlsx"))

    fj_3_name = f'{Contract_Name}第{Reporting_Periods}期进度款单位工程审核汇总对比表（附件三）.xlsx'
    save_path = os.path.join(output_path, fj_3_name)
    fj_4_name = f'{Contract_Name}第{Reporting_Periods}期进度款单位工程审核明细对比表（附件四）.xlsx'

    # 获取分部分项的文件路径

    wb_fj3 = Workbook()
    ws_fj3 = wb_fj3.active
    ws_fj3.title = "附件3"
    ws_fj3['A1'] = '附件3'
    ws_fj3['A2'] = '中国雄安集团生态建设投资有限公司'
    ws_fj3['A3'] = f'单位工程审核汇总对比表 （第 {Reporting_Periods} 次进度付款）'
    ws_fj3['A4'] = f'合同名称：{Contract_Name}' + " " * 30 + f"合同编号：{Contract_Number}" + " " * 30 + f"月份：{year}年第{month}月"
    title = ['序号', '项目名称', '合同金额（元）', '至上期累计计量金额（元）', '本期申请计量金额（元）',
             '本期审定计量金额（元）', '累计计量金额（元）', '本次核增（－）/减金额（＋）（元）']
    ws_fj3.append(title)

    # 合并单元格
    ws_fj3.merge_cells('A2:H2')
    ws_fj3.merge_cells('A3:H3')
    ws_fj3.merge_cells('A4:H4')

    # 设置列宽,从A列到I列，列宽分别为8.8，30，17.8，17.8，17.8，17.8，17.8，17.8
    ws_fj3.column_dimensions['A'].width = 8.8
    ws_fj3.column_dimensions['B'].width = 30
    for i in range(3, 9):
        ws_fj3.column_dimensions[get_column_letter(i)].width = 17.8
    # 设置所有行高为24
    for i in range(1, 100):
        ws_fj3.row_dimensions[i].height = 24
    # 设置第5行行高为40
    ws_fj3.row_dimensions[5].height = 40
    # 设置所有单元格的字体为宋体，大小为9，居中，可以换行
    for row in ws_fj3.iter_rows():
        for cell in row:
            cell.font = Font(name='宋体', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 设置第五行有边框线，设置为所有框线
    for cell in ws_fj3[5]:
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
    # 设置第4行左对齐
    for cell in ws_fj3[4]:
        cell.alignment = Alignment(horizontal='left', vertical='center')
    # 设置第2、3行字体为黑体，字号为16号
    for cell in ws_fj3[2]:
        cell.font = Font(name='黑体', size=16)
    for cell in ws_fj3[3]:
        cell.font = Font(name='黑体', size=16)
    # 写入数据
    dx_input_folder_ht = os.path.join(input_path, "ht", "ht_dx.xlsx")
    dx_input_folder_ss = os.path.join(input_path, "ss", "ss_dx.xlsx")
    dx_input_folder_sd = os.path.join(input_path, "sd", "sd_dx.xlsx")

    wb_dx_ht = pl.load_workbook(dx_input_folder_ht, data_only=True)
    ws_dx_ht = wb_dx_ht.active
    wb_dx_ss = pl.load_workbook(dx_input_folder_ss, data_only=True)
    ws_dx_ss = wb_dx_ss.active
    wb_dx_sd = pl.load_workbook(dx_input_folder_sd, data_only=True)
    ws_dx_sd = wb_dx_sd.active

    for row_ht, row_ss, row_sd in zip(ws_dx_ht.iter_rows(), ws_dx_ss.iter_rows(), ws_dx_sd.iter_rows()):
        num_ht = row_ht[column_index_from_string("A") - 1].value
        name_ht = row_ht[column_index_from_string("B") - 1].value
        price_ht = row_ht[column_index_from_string("C") - 1].value
        price_shangqi = 0
        price_ss = row_ss[column_index_from_string("C") - 1].value
        price_sd = row_sd[column_index_from_string("C") - 1].value
        leiji_price = float(price_sd) + float(price_shangqi)
        benci_price = float(price_ss) - float(price_sd)
        row_text = [num_ht, name_ht, price_ht, price_shangqi, price_ss, price_sd, leiji_price, benci_price]
        ws_fj3.append(row_text)
    max_row = ws_fj3.max_row
    ht_hj = 0
    shangqi_hj = 0
    ss_hj = 0
    sd_hj = 0
    leiji_hj = 0
    benci_hj = 0

    for row in ws_fj3.iter_rows(min_row=6, max_row=max_row,min_col=3, max_col=column_index_from_string("H")):
        for cell in row:
            cell.value = float(cell.value)
        ht_hj += row[0].value
        shangqi_hj += row[1].value
        ss_hj += row[2].value
        sd_hj += row[3].value
        leiji_hj += row[4].value
        benci_hj += row[5].value


    ws_fj3.append([None, "合计", ht_hj, shangqi_hj, ss_hj, sd_hj, leiji_hj, benci_hj])
    # 从第6行开始，到最后一行，设置所有单元格的字体为宋体，大小为9，居中，可以换行，设置所有框线
    for row in ws_fj3.iter_rows(min_row=6, max_row=max_row + 1):
        for cell in row:
            cell.font = Font(name='宋体', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    wb_fj4 = pl.load_workbook(os.path.join(output_path, fj_4_name))
    ws_fj4 = wb_fj4.active
    ws_fj4[f'G{ws_fj4.max_row}'] = ht_hj
    wb_fj4.save(os.path.join(output_path, fj_4_name))

    wb_fj3.save(save_path)
    return save_path

def create_fujian2(output_path, Contract_Name, Contract_Number, Contract_Amount, year, month, Reporting_Periods,
                    Advance_Payment_Payment_Ratio,  # 预付款支付比例
                    Advance_Payment_Deduction_Ratio, # 预付款扣回比例
                   # 截止上期部分
                    Completed_Contract_Internal_Projects_Until_Last_Period, # 截止上期累计计量金额（元）
                    Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period,  # 预付安全文明施工费
                    Accumulated_Completed_Amount_Until_Last_Period_Change_Projects, # 变更项目
                    Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project,      #计日工项目
                    Accumulated_Completed_Amount_Until_Last_Period_Claim_Project, # 索赔项目
                    Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment, # 价格调整
                    Accumulated_Completed_Amount_Until_Last_Period_Other,
                    Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount,
                    Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction,
                    Total_Completed_Output_Value,
                    Deduction_of_Supply_of_Plants_by_Party_A,
                    Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
                    Accounts_Payable_Advance_Payment,
                    Advance_Payment_Offset,
                    Temporary_Withholding_Payment_Retention_Money,
                    Refund_of_Payment_Retention_Money,  # 返还支付保留金
                    Total_Amount_Receivable,
                    Deduction_Penalty,
                    Deduction_Fine,
                    Deduction_Other,
                    Total_Deductions,
                    Total_Accounts_Payable,

                    # 本次送审的参数
                    Current_Application_Internal_Contract_Project,  # 本期申请计量金额（元）
                    Current_Application_Prepayment_Safety_Civilized_Construction_Fee,   # 预付安全文明施工费
                    Current_Application_Change_Project, # 变更项目
                    Current_Application_Time_and_Materials_Project, # 计日工项目
                    Current_Application_Claim_Project,  # 索赔项目
                    Current_Application_Price_Adjustment,
                    Current_Application_Other,
                    Current_Application_Tax_Adjustment_Payable_Amount,
                    Current_Application_Financial_Evaluation_Reduction,
                    Current_Application_Total_Completed_Output_Value,
                    Current_Application_Deduction_of_Supply_of_Plants_by_Party_A,
                    Current_Application_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
                    Current_Application_Accounts_Payable_Advance_Payment,
                    Current_Application_Advance_Payment_Offset,
                    Current_Application_Temporary_Withholding_Payment_Retention_Money,
                    Current_Application_Refund_of_Payment_Retention_Money,
                    Current_Application_Total_Amount_Receivable,
                    Current_Application_Deduction_Penalty,
                    Current_Application_Deduction_Fine,
                    Current_Application_Deduction_Other,
                    Current_Application_Total_Deductions,
                    Current_Application_Total_Accounts_Payable,
                    # 本次审核参数
                    Current_Approved_Internal_Contract_Project, # 本期审定计量金额（从附件3 中提取）
                    Current_Approved_Prepayment_Safety_Civilized_Construction_Fee,
                    Current_Approved_Change_Project,
                    Current_Approved_Time_and_Materials_Project,
                    Current_Approved_Claim_Project,
                    Current_Approved_Price_Adjustment,
                    Current_Approved_Other,
                    Current_Approved_Tax_Adjustment_Payable_Amount,
                    Current_Approved_Financial_Evaluation_Approval,
                    Current_Approved_Total_Completed_Output_Value,
                    Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A,
                    Current_Approved_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A,
                    Current_Approved_Accounts_Payable_Advance_Payment,
                    Current_Approved_Advance_Payment_Offset,
                    Current_Approved_Temporary_Withholding_Payment_Retention_Money,
                    Current_Approved_Refund_of_Payment_Retention_Money,
                    Current_Approved_Total_Amount_Receivable,
                    Current_Approved_Deduction_Penalty,
                    Current_Approved_Deduction_Fine,
                    Current_Approved_Deduction_Other,
                    Current_Approved_Total_Deductions,
                    Current_Approved_Total_Accounts_Payable,
                    ):
    fj_2_name = f'{Contract_Name}第{Reporting_Periods}期进度款工程付款审核表（附件二）.xlsx'
    save_path = os.path.join(output_path, fj_2_name)
    fj_3_name = f'{Contract_Name}第{Reporting_Periods}期进度款单位工程审核汇总对比表（附件三）.xlsx'

    fujian3_path = os.path.join(output_path, fj_3_name)
    wb_fujian3 = pl.load_workbook(fujian3_path, data_only=True)
    ws_fujian3 = wb_fujian3.active
    wb_fujian2 = Workbook()
    ws_fujian2 = wb_fujian2.active
    ws_fujian2.title = "附件2"
    ws_fujian2['A1'] = '附件2'
    ws_fujian2['A2'] = '中国雄安集团生态建设投资有限公司'
    ws_fujian2['A3'] = f'工程付款审核表 （第 {Reporting_Periods} 次进度付款）'
    ws_fujian2['A4'] = f'合同名称： + {Contract_Name}' + " " * 30 + f"合同编号：{Contract_Number}" + " " * 30 + f"月份：{year}年第{month}月"

    # 合并B6:B14单元格
    ws_fujian2.merge_cells('A2:I2')
    ws_fujian2.merge_cells('A3:I3')
    ws_fujian2.merge_cells('A4:I4')
    ws_fujian2.merge_cells('B6:B14')
    # 合并b5:c5单元格
    ws_fujian2.merge_cells('B5:C5')
    ws_fujian2.merge_cells('B15:C15')
    ws_fujian2.merge_cells('B16:C16')
    ws_fujian2.merge_cells('B17:C17')
    ws_fujian2.merge_cells('B18:B19')
    ws_fujian2.merge_cells('B20:B21')
    ws_fujian2.merge_cells('B22:C22')
    ws_fujian2.merge_cells('B23:B25')
    ws_fujian2.merge_cells('B26:C26')
    ws_fujian2.merge_cells('B27:C27')

    ws_fujian2['A5'] = '序号'
    ws_fujian2['B5'] = '项目名称'
    ws_fujian2['D5'] = '截止上期末累计完成额（元）'
    ws_fujian2['E5'] = '本期申请金额（元）'
    ws_fujian2['F5'] = '本期审定金额（元）'
    ws_fujian2['G5'] = '审核差值（“+”为减，“-”为增）'
    ws_fujian2['H5'] = '截止本期末累计审定额（元）'
    ws_fujian2['I5'] = '备注'

    ws_fujian2['A6'] = 1
    ws_fujian2['A7'] = 2
    ws_fujian2['A8'] = 3
    ws_fujian2['A9'] = 4
    ws_fujian2['A10'] = 5
    ws_fujian2['A11'] = 6
    ws_fujian2['A12'] = 7
    ws_fujian2['A13'] = 8
    ws_fujian2['A14'] = 9
    ws_fujian2['B6'] = '完成产值'
    # 合同内项目
    ws_fujian2['C6'] = '合同内项目'
    ws_fujian2['D6'] = float(Completed_Contract_Internal_Projects_Until_Last_Period)
    ws_fujian2['E6'] = float(ws_fujian3[f'E{ws_fujian3.max_row}'].value)
    ws_fujian2['F6'] = float(ws_fujian3[f'F{ws_fujian3.max_row}'].value)

    # 预付安全文明施工费
    ws_fujian2['C7'] = '预付安全文明施工费'
    ws_fujian2['D7'] = float(Completed_Prepaid_Safety_Civilized_Construction_Fee_Until_Last_Period)
    ws_fujian2['E7'] = float(Current_Application_Prepayment_Safety_Civilized_Construction_Fee)
    ws_fujian2['F7'] = float(Current_Approved_Prepayment_Safety_Civilized_Construction_Fee)
    # 变更项目
    ws_fujian2['C8'] = '变更项目'
    ws_fujian2['D8'] = float(Accumulated_Completed_Amount_Until_Last_Period_Change_Projects)
    ws_fujian2['E8'] = float(Current_Application_Change_Project)
    ws_fujian2['F8'] = float(Current_Approved_Change_Project)
    ws_fujian2['C9'] = '计日工项目'
    ws_fujian2['D9'] = float(Accumulated_Completed_Amount_Until_Last_Period_Time_and_Materials_Project)
    ws_fujian2['E9'] = float(Current_Application_Time_and_Materials_Project)
    ws_fujian2['F9'] = float(Current_Approved_Time_and_Materials_Project)

    ws_fujian2['C10'] = '索赔项目'
    ws_fujian2['D10'] = float(Accumulated_Completed_Amount_Until_Last_Period_Claim_Project)
    ws_fujian2['E10'] = float(Current_Application_Claim_Project)
    ws_fujian2['F10'] = float(Current_Approved_Claim_Project)



    ws_fujian2['C11'] = '价格调整'
    ws_fujian2['D11'] = float(Accumulated_Completed_Amount_Until_Last_Period_Price_Adjustment)
    ws_fujian2['E11'] = float(Current_Application_Price_Adjustment)
    ws_fujian2['F11'] = float(Current_Approved_Price_Adjustment)


    ws_fujian2['C12'] = '其他'
    ws_fujian2['D12'] = float(Accumulated_Completed_Amount_Until_Last_Period_Other)
    ws_fujian2['E12'] = float(Current_Application_Other)
    ws_fujian2['F12'] = float(Current_Approved_Other)



    ws_fujian2['C13'] = '税率调整应付金额'
    ws_fujian2['D13'] = float(Accumulated_Completed_Amount_Until_Last_Period_Tax_Adjustment_Payable_Amount)
    ws_fujian2['E13'] = float(Current_Application_Tax_Adjustment_Payable_Amount)
    ws_fujian2['F13'] = float(Current_Approved_Tax_Adjustment_Payable_Amount)



    ws_fujian2['C14'] = '财评审减'
    ws_fujian2['D14'] = float(Accumulated_Completed_Amount_Until_Last_Period_Financial_Evaluation_Reduction)
    ws_fujian2['E14'] = float(Current_Application_Financial_Evaluation_Reduction)
    ws_fujian2['F14'] = float(Current_Approved_Financial_Evaluation_Approval)



    ws_fujian2['A15'] = 10
    ws_fujian2['B15'] = '完成产值合计（9=1+2+3+4+5+6+7+8）'
    ws_fujian2['D15'] = float(Total_Completed_Output_Value)
    ws_fujian2['E15'] = float(Current_Application_Total_Completed_Output_Value)
    ws_fujian2['F15'] = float(Current_Approved_Total_Completed_Output_Value)

    print(f"审定完成产值合计:{Current_Approved_Total_Completed_Output_Value}")

    for row in ws_fujian2.iter_rows():
        if row[0].row < 6:
            continue
        write_row = row[0].row
        if ws_fujian2[f'E{write_row}'].value is None or ws_fujian2[f'E{write_row}'].value == "":
            ws_fujian2[f'E{write_row}'].value = 0
        if ws_fujian2[f'F{write_row}'].value is None or ws_fujian2[f'F{write_row}'].value == "":
            ws_fujian2[f'F{write_row}'].value = 0
        if ws_fujian2[f'D{write_row}'].value is None or ws_fujian2[f'D{write_row}'].value == "":
            ws_fujian2[f'D{write_row}'].value = 0
        ws_fujian2[f'G{write_row}'] = ws_fujian2[f'E{write_row}'].value - ws_fujian2[f'F{write_row}'].value
        ws_fujian2[f'H{write_row}'] = ws_fujian2[f'D{write_row}'].value + ws_fujian2[f'F{write_row}'].value

    ws_fujian2['A16'] = 11
    ws_fujian2['B16'] = '扣减甲供材料'
    ws_fujian2['D16'] = float(Deduction_of_Supply_of_Plants_by_Party_A)
    ws_fujian2['E16'] = float(Current_Application_Deduction_of_Supply_of_Plants_by_Party_A)
    ws_fujian2['F16'] = float(Current_Approved_Deduction_of_Supply_of_Plants_by_Party_A)

    ws_fujian2['A17'] = 12
    ws_fujian2['B17'] = '扣除甲供苗后完成产值合计'
    ws_fujian2['D17'] = float(Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A)
    ws_fujian2['E17'] = float(Current_Application_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A)
    ws_fujian2['F17'] = float(Current_Approved_Total_Completed_Output_Value_After_Deducting_Supply_of_Plants_by_Party_A)

    ws_fujian2['A18'] = 13
    ws_fujian2['A19'] = 14
    ws_fujian2['B18'] = '预付款'
    ws_fujian2['C18'] = '应付预付款'
    ws_fujian2['D18'] = float(Accounts_Payable_Advance_Payment)
    ws_fujian2['E18'] = float(Current_Application_Accounts_Payable_Advance_Payment)
    ws_fujian2['F18'] = float(Current_Approved_Accounts_Payable_Advance_Payment)

    ws_fujian2['C19'] = '预付款抵扣'
    ws_fujian2['D19'] = float(Advance_Payment_Offset)
    ws_fujian2['E19'] = float(Current_Application_Advance_Payment_Offset)
    Advance_Payment_Payment_Ratio = float(Advance_Payment_Payment_Ratio)/100
    Advance_Payment_Deduction_Ratio = float(Advance_Payment_Deduction_Ratio)/100
    # （计量与计价金额累计-签约合同价*30%）/(签约合同价*80%-签约合同价*30%)*已付预付款额-上期累计扣回预付款金额
    price_leiji = ws_fujian2['H15'].value
    leiji_amount = round((float(price_leiji) / float(Contract_Amount)) * 100, 2)
    print(price_leiji, leiji_amount)
    if leiji_amount < 30:
        ws_fujian2['F19'] = 0
    elif 30 <= leiji_amount < 80:
        ws_fujian2['F19'] = round((ws_fujian2['H15'].value - float(Contract_Amount) * Advance_Payment_Payment_Ratio) / (float(Contract_Amount) * Advance_Payment_Deduction_Ratio - float(Contract_Amount) * Advance_Payment_Payment_Ratio) * ws_fujian2['D18'].value - ws_fujian2['D19'].value,2)
    ws_fujian2['A20'] = 15
    ws_fujian2['A21'] = 16
    ws_fujian2['B20'] = '支付保留金'
    ws_fujian2['C20'] = '暂扣支付保留金'
    ws_fujian2['D20'] = float(Temporary_Withholding_Payment_Retention_Money)
    ws_fujian2['E20'] = float(Current_Application_Temporary_Withholding_Payment_Retention_Money)
    ws_fujian2['F20'] = float(Current_Approved_Temporary_Withholding_Payment_Retention_Money)

    ws_fujian2['C21'] = '返还支付保留金'
    ws_fujian2['D21'] = float(Refund_of_Payment_Retention_Money)
    ws_fujian2['E21'] = float(Current_Application_Refund_of_Payment_Retention_Money)
    ws_fujian2['F21'] = float(Current_Approved_Refund_of_Payment_Retention_Money)


    ws_fujian2['A22'] = 17
    ws_fujian2['B22'] = '应得款合计（17=12+13-14-15+16）'
    ws_fujian2['D22'] = float(Total_Amount_Receivable)
    ws_fujian2['E22'] = float(Current_Application_Total_Amount_Receivable)
    ws_fujian2['F22'] = float(Current_Approved_Total_Amount_Receivable)

    ws_fujian2['A23'] = 18
    ws_fujian2['A24'] = 19
    ws_fujian2['A25'] = 20
    ws_fujian2['B23'] = '应扣款'
    ws_fujian2['C23'] = '违约金'
    ws_fujian2['D23'] = float(Deduction_Penalty)
    ws_fujian2['E23'] = float(Current_Application_Deduction_Penalty)
    ws_fujian2['F23'] = float(Current_Approved_Deduction_Penalty)


    ws_fujian2['C24'] = '罚款'
    ws_fujian2['D24'] = float(Deduction_Fine)
    ws_fujian2['E24'] = float(Current_Application_Deduction_Fine)
    ws_fujian2['F24'] = float(Current_Approved_Deduction_Fine)
    ws_fujian2['C25'] = '其他'
    ws_fujian2['D25'] = float(Deduction_Other)
    ws_fujian2['E25'] = float(Current_Application_Deduction_Other)
    ws_fujian2['F25'] = float(Current_Approved_Deduction_Other)

    ws_fujian2['A26'] = 21
    ws_fujian2['B26'] = '应扣款合计（21=18+19+20）'
    ws_fujian2['D26'] = float(Total_Deductions)
    ws_fujian2['E26'] = float(Current_Application_Total_Deductions)
    print(Current_Application_Total_Deductions)
    ws_fujian2['F26'] = float(Current_Approved_Total_Deductions)

    ws_fujian2['A27'] = 22
    ws_fujian2['B27'] = '应付款合计（22=17-21）'
    ws_fujian2['D27'] = float(Total_Accounts_Payable)
    ws_fujian2['E27'] = float(Current_Application_Total_Accounts_Payable)
    ws_fujian2['F27'] = float(Current_Approved_Total_Accounts_Payable)
    print(Current_Approved_Total_Accounts_Payable)

    for row in ws_fujian2.iter_rows():
        if row[0].row < 6:
            continue
        write_row = row[0].row
        ws_fujian2[f'G{write_row}'] = ws_fujian2[f'E{write_row}'].value - ws_fujian2[f'F{write_row}'].value
        ws_fujian2[f'H{write_row}'] = ws_fujian2[f'D{write_row}'].value + ws_fujian2[f'F{write_row}'].value

    # 设置所有单元格的字体为宋体，大小为9，居中，可以换行，设置所有框线
    for row in ws_fujian2.iter_rows(min_row=5):
        for cell in row:
            cell.font = Font(name='宋体', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    # 第1、2列列宽设置为5，其余设置为15
    ws_fujian2.column_dimensions['A'].width = 5
    ws_fujian2.column_dimensions['B'].width = 5
    for i in range(3, 10):
        ws_fujian2.column_dimensions[get_column_letter(i)].width = 15

    # 第2、3行行高设置为40，其余设置为24
    ws_fujian2.row_dimensions[2].height = 40
    ws_fujian2.row_dimensions[3].height = 40
    for i in range(4, 28):
        ws_fujian2.row_dimensions[i].height = 24

    # 设置第2、3行字体为黑体，字号为16号
    for cell in ws_fujian2[2]:
        cell.font = Font(name='黑体', size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in ws_fujian2[3]:
        cell.font = Font(name='黑体', size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 设置第1行字体为黑体，字号为12号
    for cell in ws_fujian2[1]:
        cell.font = Font(name='黑体', size=12)

    # 第15、16、17、22、26、27行设置加粗
    for number in [15, 16, 17, 22, 26, 27]:
        for cell in ws_fujian2[number]:
            cell.font = Font(bold=True, size=9)

    wb_fujian2.save(save_path)
    # os.system(f"open {save_path}")
    return save_path

def create_fujian1(output_path, Contract_Name, Contract_Number, year, month, Reporting_Periods, Construction_Unit,
                   Contract_Amount, Funding_Source, Consulting_Unit):
    fj_1_name = f'{Contract_Name}第{Reporting_Periods}期进度款付款汇总表（附件一）.xlsx'
    save_path = os.path.join(output_path, fj_1_name)
    wb_fujian1 = Workbook()
    ws_fujian1 = wb_fujian1.active
    ws_fujian1.title = "附件1"
    ws_fujian1['A1'] = '附件1'
    ws_fujian1['A2'] = f'中国雄安集团生态建设投资有限公司\n付款汇总表 （第 {Reporting_Periods} 期进度付款）'
    ws_fujian1['A3'] = '合同名称'
    ws_fujian1['D3'] = '合同编号'

    ws_fujian1['A4'] = '申请单位名称'
    ws_fujian1['A5'] = '【A】原合同金额（元）'
    ws_fujian1['B5'] = '小写'
    ws_fujian1['B6'] = '大写'
    ws_fujian1['A7'] = '【B】补充协议调整后合同金额（元）'
    ws_fujian1['B7'] = '小写'
    ws_fujian1['B8'] = '大写'
    ws_fujian1['A9'] = "【C'】上期末累计核定产值（元）"
    ws_fujian1['D9'] = "【D'】本期末累计核定产值（元）"
    ws_fujian1['A10'] = "【E'】本期核定产值（元）（E=D-C）"
    ws_fujian1['D10'] = "【F'】本期末累计核定产值占合同比例/%（F=D/B)"
    ws_fujian1['A11'] = "【C】上期末累计核定产值（扣除甲供苗）（元）"
    ws_fujian1['D11'] = "【D】本期末累计核定产值（扣除甲供苗）（元）"
    ws_fujian1['A12'] = "【E】本期核定产值（扣除甲供苗）（元）（E=D-C）"
    ws_fujian1['D12'] = "【F】本期末累计核定产值占合同比例/%（F=D/B)"
    ws_fujian1['A13'] = "【G】上期末累计应得款（元）"
    ws_fujian1['D13'] = "【H】本期末累计应得款（元）"
    ws_fujian1['A14'] = "【I】本期应得款（元）（I=H-G）"
    ws_fujian1['D14'] = "【J】本期末累计应得款占合同比例/%（J=H/B)"
    ws_fujian1['A15'] = "【K】待抵扣预付款余额（元）"
    ws_fujian1['D15'] = "【L】待返还支付保留金余额（元）"
    ws_fujian1['A16'] = "【M】本期其他扣款（元）"
    ws_fujian1['D16'] = "【N】本期末累计其他扣款（罚款等）（元）"
    ws_fujian1['A17'] = "【O】本期应付款（小写）（元）（O=I-M）"
    ws_fujian1['A18'] = "本期应付款（大写）（元）"
    ws_fujian1['A19'] = "审核情况说明（相关款项相见工程付款审核表）"

    ws_fujian1['A21'] = "审查人（资格印章或签字）："
    ws_fujian1['A22'] = "复核人（资格印章或签字）："
    ws_fujian1['D22'] = f"编制单位：{Consulting_Unit}\n（盖章）"
    # 根据表格内容，开始合并
    ws_fujian1.merge_cells('A2:F2')
    ws_fujian1.merge_cells('B3:C3')
    ws_fujian1.merge_cells('E3:F3')
    ws_fujian1.merge_cells('B4:F4')
    ws_fujian1.merge_cells('A5:A6')
    ws_fujian1.merge_cells('C5:F5')
    ws_fujian1.merge_cells('C6:F6')
    ws_fujian1.merge_cells('A7:A8')
    ws_fujian1.merge_cells('C7:F7')
    ws_fujian1.merge_cells('C8:F8')
    for row in ws_fujian1.iter_rows(min_row=9, max_row=16):
        ws_fujian1.merge_cells(f'A{row[0].row}:B{row[0].row}')
        ws_fujian1.merge_cells(f'D{row[0].row}:E{row[0].row}')

    ws_fujian1.merge_cells('A17:B17')
    ws_fujian1.merge_cells('C17:F17')

    ws_fujian1.merge_cells('A18:B18')
    ws_fujian1.merge_cells('C18:F18')

    ws_fujian1.merge_cells('A19:F19')
    ws_fujian1.merge_cells('A20:F20')
    ws_fujian1.merge_cells('A21:C21')
    ws_fujian1.merge_cells('A22:C22')
    ws_fujian1.merge_cells('D22:F22')

    # 设置字体为宋体，字号9号，居中，可以换行，从第3行至第18行设置所有框线
    for row in ws_fujian1.iter_rows(min_row=3, max_row=18):
        for cell in row:
            cell.font = Font(name='宋体', size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))


    # 设置A列宽15， B列宽10， C列宽20， D列宽15， E列宽10， F列宽20
    ws_fujian1.column_dimensions['A'].width = 15
    ws_fujian1.column_dimensions['B'].width = 10
    ws_fujian1.column_dimensions['C'].width = 20
    ws_fujian1.column_dimensions['D'].width = 15
    ws_fujian1.column_dimensions['E'].width = 10
    ws_fujian1.column_dimensions['F'].width = 20

    # 设置第2行行高为40，第3行行高为24，其余行高为20
    ws_fujian1.row_dimensions[2].height = 40
    ws_fujian1.row_dimensions[3].height = 24
    for i in range(4, 23):
        ws_fujian1.row_dimensions[i].height = 40
    # 设置第20行行高为60
    ws_fujian1.row_dimensions[20].height = 60
    # 设置第2行字体为黑体，字号为16号
    for cell in ws_fujian1[2]:
        cell.font = Font(name='黑体', size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 设置第3行至最后一行单元格可换行
    for row in ws_fujian1.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 写入变量
    ws_fujian1['B3'] = Contract_Name
    ws_fujian1['E3'] = Contract_Number
    ws_fujian1['B4'] = Construction_Unit
    ws_fujian1['C5'] = Contract_Amount
    ws_fujian1['C6'] = number_to_chinese(float(Contract_Amount))
    ws_fujian1['C7'] = float(ws_fujian1['C5'].value)
    ws_fujian1['C8'] = ws_fujian1['C6'].value
    # import wb_fujian2
    fj_2_name = f'{Contract_Name}第{Reporting_Periods}期进度款工程付款审核表（附件二）.xlsx'
    wb_fujian2 = pl.load_workbook(os.path.join(output_path, fj_2_name), data_only=True)
    ws_fujian2 = wb_fujian2.active
    for row in ws_fujian2.iter_rows():
        for cell in row:
            if cell is None or cell == "":
                cell.value = 0
                print(cell.value)
    ws_fujian1['C9'] = ws_fujian2['D15'].value
    ws_fujian1['F9'] = ws_fujian2['H15'].value
    print(f"错误检查：{ws_fujian1['C9'].value, ws_fujian1['F9'].value}")

    ws_fujian1['C10'] = ws_fujian1['F9'].value - ws_fujian1['C9'].value
    print(ws_fujian1['C10'].value, type(ws_fujian1['C10'].value), ws_fujian1['C7'].value, type(ws_fujian1['C7'].value))

    ws_fujian1['F10'] = ws_fujian1['F9'].value / ws_fujian1['C7'].value
    # 设置F10单元格以%百分比形式显示，保留两位小数
    ws_fujian1['F10'].number_format = '0.00%'

    ws_fujian1['C11'] = ws_fujian2['D17'].value
    ws_fujian1['F11'] = ws_fujian2['H17'].value
    print(ws_fujian1['C11'].value, ws_fujian1['F11'].value)
    if ws_fujian1['C11'].value is None or ws_fujian1['F11'].value is None:
        ws_fujian1['C12'].value = 0
        ws_fujian1['F12'].value = 0
    else:
        ws_fujian1['C12'] = ws_fujian1['F11'].value - ws_fujian1['C11'].value
        ws_fujian1['F12'] = ws_fujian1['F11'].value / ws_fujian1['C7'].value
    # 设置F12单元格以%百分比形式显示，保留两位小数
    ws_fujian1['F12'].number_format = '0.00%'

    ws_fujian1['C13'] = ws_fujian2['D22'].value
    ws_fujian1['F13'] = ws_fujian2['H22'].value
    if ws_fujian1['C13'].value is None or ws_fujian1['F13'].value is None:
        ws_fujian1['C14'].value = 0
        ws_fujian1['F14'].value = 0
    else:
        ws_fujian1['C14'] = round(ws_fujian1['F13'].value - ws_fujian1['C13'].value, 2)
        ws_fujian1['F14'] = ws_fujian1['F13'].value / ws_fujian1['C7'].value
    # 设置F14单元格以%百分比形式显示，保留两位小数
    ws_fujian1['F14'].number_format = '0.00%'

    print(ws_fujian2['H18'].value, ws_fujian2['H19'].value)
    if ws_fujian2['H18'].value is None or ws_fujian2['H19'].value is None:
        ws_fujian1['C15'].value = 0
    else:
        ws_fujian1['C15'] = round(ws_fujian2['H18'].value - ws_fujian2['H19'].value, 2)
    ws_fujian1['F15'] = ws_fujian2['H20'].value
    ws_fujian1['C16'] = ws_fujian2['F26'].value
    ws_fujian1['F16'] = ws_fujian2['H26'].value
    if ws_fujian1['C16'].value is None:
        ws_fujian1['C16'].value = 0
    if ws_fujian1['F16'].value is None:
        ws_fujian1['F16'].value = 0
    Accounts_Payable_Advance_Payment = round(ws_fujian1['C14'].value - ws_fujian1['C16'].value,2)
    print(Accounts_Payable_Advance_Payment)
    ws_fujian1['C17'] = Accounts_Payable_Advance_Payment
    ws_fujian1['C18'] = number_to_chinese(float(Accounts_Payable_Advance_Payment))

    ws_fujian1['A20'] = (
        f"该工程资金来源为：{Funding_Source}，由{Construction_Unit}施工，于{year}年{month}月报送经监理单位确认的《工程款支付申请表》，"
        f"报送工程造价为{round(ws_fujian2['E6'].value, 2)}元（含甲供材料），我公司依据建设工程施工合同、招标文件及投标文件，并根据形象进度，经过实际核算，"
        f"最终审核确定完成工程造价为{ws_fujian2['F6'].value}元（含甲供材料），核减金额为{round(ws_fujian2['G6'].value, 2)}元，应付款金额为{ws_fujian2['F27'].value}元。"
        f"注：计量审核仅作为施工进度的拨款依据。")

    ws_fujian1['C9'].number_format = '0.00'
    ws_fujian1['F9'].number_format = '0.00'
    ws_fujian1['C10'].number_format = '0.00'
    ws_fujian1['C11'].number_format = '0.00'
    ws_fujian1['F11'].number_format = '0.00'
    ws_fujian1['C13'].number_format = '0.00'
    ws_fujian1['F13'].number_format = '0.00'
    ws_fujian1['C14'].number_format = '0.00'
    ws_fujian1['C15'].number_format = '0.00'
    ws_fujian1['F15'].number_format = '0.00'
    ws_fujian1['C16'].number_format = '0.00'
    ws_fujian1['F16'].number_format = '0.00'
    ws_fujian1['C17'].number_format = '0.00'
    # ws_fujian1['C16'].number_format = '0.00'
    # ws_fujian1['C16'].number_format = '0.00'
    # ws_fujian1['C16'].number_format = '0.00'


    # 加粗
    wb_fujian1.save(save_path)
    return save_path


# 创建一个Excel表格，用于保存新项目数据，重复利用
def creat_info_project_excel(Contract_Name, output_path, Contract_Number, Principal_Party, Construction_Unit, Supervision_Unit,
                             Design_Unit, Consulting_Unit, Project_Overview, Funding_Source, Price_Form, Progress_Payment_Ratio, Contract_Amount,Advance_Payment_Payment_Ratio,Advance_Payment_Deduction_Ratio):
    project_info_file_name = "project_info.xlsx"
    save_path = os.path.join(output_path, project_info_file_name)
    # 尝试连接到名称为project_info.xlsx的文件，如果存在，则打开，否则新建一个
    project_info_dict = {}
    try:
        wb = pl.load_workbook(save_path, data_only=True)
        print(f'文件已存在，打开')
        print(wb.sheetnames)
        if Contract_Name in wb.sheetnames:
            ws = wb[Contract_Name]
            # 清空这个sheet表格，倒叙删除所有行
            for i in range(ws.max_row, 0, -1):
                ws.delete_rows(i)
            title = ["name_in_chinese", "name_in_english", "value"]
            ws.append(title)
            ws.append(['合同名称', "Contract_Name", Contract_Name])
            ws.append(['合同编号', "Contract_Number", Contract_Number])
            ws.append(['建设单位', "Principal_Party", Principal_Party])
            ws.append(['施工单位', "Construction_Unit", Construction_Unit])
            ws.append(['监理单位', "Supervision_Unit", Supervision_Unit])
            ws.append(['设计单位', "Design_Unit", Design_Unit])
            ws.append(['咨询单位', "Consulting_Unit", Consulting_Unit])
            ws.append(['工程概况', "Project_Overview", Project_Overview])
            ws.append(['资金来源', "Funding_Source", Funding_Source])
            ws.append(['合同单价形式', "Price_Form", Price_Form])
            ws.append(['合同支付比例', "Progress_Payment_Ratio", Progress_Payment_Ratio])
            ws.append(['合同金额', "Contract_Amount", Contract_Amount])
            ws.append(['预付款支付比例', "Advance_Payment_Payment_Ratio", Advance_Payment_Payment_Ratio])
            ws.append(['预付款抵扣比例', "Advance_Payment_Deduction_Ratio", Advance_Payment_Deduction_Ratio])
            for row in ws.iter_rows():
                project_info_dict[row[1].value] = row[2].value
            wb.save(save_path)
            return project_info_dict
        else:
            print(f'新建了一个sheet: {Contract_Name}')
            ws = wb.create_sheet(Contract_Name)
            title = ["name_in_chinese", "name_in_english", "value"]
            ws.append(title)
            ws.append(['合同名称', "Contract_Name", Contract_Name])
            ws.append(['合同编号', "Contract_Number", Contract_Number])
            ws.append(['建设单位', "Principal_Party", Principal_Party])
            ws.append(['施工单位', "Construction_Unit", Construction_Unit])
            ws.append(['监理单位', "Supervision_Unit", Supervision_Unit])
            ws.append(['设计单位', "Design_Unit", Design_Unit])
            ws.append(['咨询单位', "Consulting_Unit", Consulting_Unit])
            ws.append(['工程概况', "Project_Overview", Project_Overview])
            ws.append(['资金来源', "Funding_Source", Funding_Source])
            ws.append(['合同单价形式', "Price_Form", Price_Form])
            ws.append(['合同支付比例', "Progress_Payment_Ratio", Progress_Payment_Ratio])
            ws.append(['合同金额', "Contract_Amount", Contract_Amount])
            ws.append(['预付款支付比例', "Advance_Payment_Payment_Ratio", Advance_Payment_Payment_Ratio])
            ws.append(['预付款抵扣比例', "Advance_Payment_Deduction_Ratio", Advance_Payment_Deduction_Ratio])
            for row in ws.iter_rows():
                project_info_dict[row[1].value] = row[2].value
            wb.save(save_path)
            return project_info_dict
    except FileNotFoundError:
        print(f'文件不存在，新建')
        wb = Workbook()
        ws = wb.create_sheet(Contract_Name)
        title = ["name_in_chinese", "name_in_english", "value"]
        ws.append(title)
        ws.append(['合同名称', "Contract_Name", Contract_Name])
        ws.append(['合同编号', "Contract_Number", Contract_Number])
        ws.append(['建设单位', "Principal_Party", Principal_Party])
        ws.append(['施工单位', "Construction_Unit", Construction_Unit])
        ws.append(['监理单位', "Supervision_Unit", Supervision_Unit])
        ws.append(['设计单位', "Design_Unit", Design_Unit])
        ws.append(['咨询单位', "Consulting_Unit", Consulting_Unit])
        ws.append(['工程概况', "Project_Overview", Project_Overview])
        ws.append(['资金来源', "Funding_Source", Funding_Source])
        ws.append(['合同单价形式', "Price_Form", Price_Form])
        ws.append(['合同支付比例', "Progress_Payment_Ratio", Progress_Payment_Ratio])
        ws.append(['合同金额', "Contract_Amount", Contract_Amount])
        ws.append(['预付款支付比例', "Advance_Payment_Payment_Ratio", Advance_Payment_Payment_Ratio])
        ws.append(['预付款抵扣比例', "Advance_Payment_Deduction_Ratio", Advance_Payment_Deduction_Ratio])
        for row in ws.iter_rows():
            project_info_dict[row[1].value] = row[2].value
        project_info_file_name = "project_info.xlsx"
        save_path = os.path.join(output_path, project_info_file_name)
        wb.save(save_path)
        return project_info_dict


def get_project_data_dict(output_path, Contract_Name):
    project_info_dict = {}
    file_path = os.path.join(output_path, "project_info.xlsx")
    wb = pl.load_workbook(file_path, data_only=True)
    ws = wb[Contract_Name]
    for row in ws.iter_rows():
        project_info_dict[row[1].value] = row[2].value
    return project_info_dict

def check_last(Contract_Name, Reporting_Periods, output_path):
    # 检查是否存在上一期的文件夹，如果存在的话，则获取其中的附件1、附件2、附件3、附件4 中上期的数据，并返回内容待用

    # 定位到项目名称的文件夹
    target_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "function", "progress_reports_xd", Contract_Name)
    print(target_dir)
    folder_name = f'第{Reporting_Periods - 1}期'

    # 检查附件1、2、3、4是否存在，如果存在，则读取上期数据，如果不存在，则pass
    fj_1_name = f'{Contract_Name}第{Reporting_Periods - 1}期进度款付款汇总表（附件一）.xlsx'
    fj_2_name = f'{Contract_Name}第{Reporting_Periods - 1}期进度款工程付款审核表（附件二）.xlsx'
    fj_3_name = f'{Contract_Name}第{Reporting_Periods - 1}期进度款单位工程审核汇总对比表（附件三）.xlsx'
    fj_4_name = f'{Contract_Name}第{Reporting_Periods - 1}期进度款单位工程审核明细对比表（附件四）.xlsx'
    if os.path.exists(os.path.join(output_path, fj_1_name)):
        print(f'检查到上期附件1，开始读取上期数据')
        # 读取附件1
        wb_fujian1 = pl.load_workbook(os.path.join(output_path, fj_1_name), data_only=True)
        print(f'附件1的sheet表格名称都有：{wb_fujian1.sheetnames}')
        # 先获取附件四，并读取附件4的H列和I列
        wb_fujian4 = pl.load_workbook(os.path.join(output_path, fj_4_name), data_only=True)
        ws_fujian4 = wb_fujian4.active
        for row in ws_fujian4.iter_rows():
            if row[0].row < 6:
                continue
            else:
                print(f'附件4的第{row[0].row}行的H列和I列内容为：{row[7].value}，{row[8].value}')
    else:
        print(f'未检查到上期文件，开始新建上期数据')

    return target_dir

def test(Contract_Name, Reporting_Periods):
    project_last_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "function", "progress_reports_xd", Contract_Name, f'第{Reporting_Periods - 1}期', "output_files")
    fj_3_name_last = f'{Contract_Name}第{Reporting_Periods - 1}期进度款单位工程审核汇总对比表（附件三）.xlsx'
    G_result = []
    print(os.listdir(project_last_dir))
    wb = pl.load_workbook(os.path.join(project_last_dir, fj_3_name_last), data_only=True)
    ws = wb.active
    for row in ws.iter_rows():
        if row[0].row < 6:
            continue
        else:
            print(f'附件3的第{row[0].row}行的G列内容为：{row[column_index_from_string("G") - 1].value}')
            G_result.append(row[column_index_from_string("G") - 1].value)
    wb_new = Workbook()
    ws_new = wb_new.active

    for i, g in enumerate(G_result):
        ws_new[f'A{i + 6}'] = g
    save_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "test.xlsx")
    wb_new.save(save_path)
    return save_path

def linshifujian4_4(Contract_Name):
    # 获取project_info.xlsx中的数据
    output_path = r'/Users/willcha/library_for_search/function/progress_reports_xd'
    project_info_dict = get_project_data_dict(output_path, Contract_Name)
    return project_info_dict
if __name__ == '__main__':
    Contract_Name = '容东民俗公园二期二标段施工总承包'
    print(linshifujian4_4(Contract_Name)['Contract_Number'])