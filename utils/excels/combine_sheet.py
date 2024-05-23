import os.path

import openpyxl as pl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import Workbook
from tkinter import filedialog
import tkinter as tk


def combine_sheet(file):
    """合并成无格式"""
    global row_list
    file_name = file.split("/")[-1]
    data_list = []
    wb = pl.load_workbook(file, data_only=True)
    ws_names = wb.sheetnames
    for name in ws_names:
        ws = wb[name]
        row_list = []
        for row in ws.iter_rows(max_row=ws.max_row):
            cell_list = []
            for cell in row:
                value_cell = cell.value
                if cell.value is None:
                    value_cell = 0
                cell_list.append(value_cell)
            row_list.append(cell_list)
        data_list.append(row_list)

    wb_new = Workbook()
    ws_new = wb_new.active
    title = ['序号', "项目编码","项目名称", None, "项目特征","计量单位","工程数量", "主材", None,"综合单价","合价"]
    ws_new.append(title)
    for row_list in data_list:
        for cell_list in row_list:
            ws_new.append(cell_list)
    file_name_combine = file_name.split(".")[0]
    wb_new.save(f"{file_name_combine} - 合并.xlsx")
    print(f'{file_name_combine}已完成合并')


def select_file():
    file_path = filedialog.askopenfilename()
    return file_path
