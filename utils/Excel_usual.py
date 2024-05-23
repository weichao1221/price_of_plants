# # -*- coding: utf-8 -*-
import os
import pandas
import xlrd
import openpyxl
from openpyxl import Workbook, load_workbook
from fastapi import UploadFile

# 创建一个将xls文件转换为xlsx文件的函数
def convert_xls_to_xlsx(xls_filename, xlsx_filename, save_path):
    # 读取xls文件
    xls_workbook = xlrd.open_workbook(xls_filename)
    xls_sheet = xls_workbook.sheet_by_index(0)

    # 创建一个新的xlsx文件
    xlsx_workbook = Workbook()
    xlsx_sheet = xlsx_workbook.active

    # 将数据从xls复制到xlsx
    for row_index in range(xls_sheet.nrows):
        # 跳过前4行
        # if row_index < 4:
        #     continue
        for col_index in range(xls_sheet.ncols):
            xlsx_sheet.cell(row=row_index + 1, column=col_index + 1, value=xls_sheet.cell_value(row_index, col_index))
    # save_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources')
    # 删除源文件
    os.remove(xls_filename)
    # 保存xlsx文件
    xlsx_workbook.save(os.path.join(save_path, xlsx_filename))
